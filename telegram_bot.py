from telegram.ext import (
    CallbackQueryHandler, ApplicationBuilder, CommandHandler,
    MessageHandler, filters, ConversationHandler
)
from zoneinfo import ZoneInfo
import regex as re
from decimal import Decimal
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
import os
import django
from asgiref.sync import sync_to_async
from django.db import models
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'project.settings')
django.setup()
from datetime import datetime, date
from app.models import CustomUser, Taminotchi, Pul_olish, Pul_berish,Harajatlar,CustomUser
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
import logging
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from datetime import date, datetime
from telegram import Bot
from asgiref.sync import sync_to_async
from django.db.models import Q

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

today = datetime.now(ZoneInfo("Asia/Tashkent")).date()
# NEW: Ishchilar states# --- Global Constants ---
PAGE_SIZE = 10
PAGINATION_SIZE = 10

# --- Login ---
ASK_LOGIN = 10

# --- Supplier / Qarz ---
CHOOSE_SUPPLIER = 20
SUPPLIER_NAME = 21
SUPPLIER_PHONE = 22
FIND_SUPPLIER = 23
SUPPLIER_ACTION = 24
ASK_AMOUNT = 25
ASK_DATE = 26
ASK_REASON = 27
CONFIRMATION = 28
PAYMENT_CURRENCY = 29

# --- Payment Flow ---
ASK_PAYMENT_CURRENCY = 30
ASK_PAYMENT_AMOUNT = 31
ASK_PAYMENT_DATE = 32
PAYMENT_CONFIRM = 33

# --- Excel ---
EXCEL_MENU = 40
EXCEL_DATE_START = 41
EXCEL_DATE_END = 42

# --- Report ---
REPORT_MENU = 50
REPORT_DATE_START = 51
REPORT_DATE_END = 52

# --- Rejalashtirilgan Tolovlar / Xabarlar ---
REJALA_SUPPLIER_LIST = 1000
REJALA_PAYMENT_LIST = 1001
REJALA_PAYMENT_ACTION = 1002
XABAR_PAYMENT_LIST = 1003
XABAR_PAYMENT_ACTION = 1004
PAYMENT_EDIT_AMOUNT = 1005
PAYMENT_EDIT_DATE = 1006

# --- Ishchilar ---
CHOOSE_ISHCHI = 1100
ISHCHI_NAME = 1101
ISHCHI_SURNAME = 1102
ISHCHI_PHONE = 1103
ISHCHI_PERMISSIONS = 1104
ISHCHI_CONFIRM = 1105
FIND_ISHCHI = 1106
ISHCHI_ACTION = 1107
ISHCHI_DELETE_CONFIRM = 1108
ISHCHI_EDIT_NAME = 1109
ISHCHI_EDIT_SURNAME = 1110
ISHCHI_EDIT_PHONE = 1111
ISHCHI_EDIT_PERMISSIONS = 1112
ISHCHI_EDIT_CONFIRM = 1113

# --- Harajatlar ---
HARAJAT_MENU = 1200
HARAJAT_TYPE = 1201
HARAJAT_AMOUNT = 1202
HARAJAT_DATE = 1203
HARAJAT_REASON = 1204
HARAJAT_CONFIRMATION = 1205
CHOOSE_ISHCHI_EXPENSE = 1206
HARAJAT_LIST = 1207
HARAJAT_ACTION = 1208
HARAJAT_DELETE_CONFIRM = 1209
HARAJAT_EDIT_AMOUNT = 1210
HARAJAT_EDIT_DATE = 1211
HARAJAT_EDIT_REASON = 1212
HARAJAT_EDIT_ISHCHI = 1213
HARAJAT_EDIT_CONFIRM = 1214
HARAJAT_EXCEL_MENU = 1215
HARAJAT_EXCEL_DATE_START = 1216
HARAJAT_EXCEL_DATE_END = 1217



scheduler = AsyncIOScheduler()

# Notification function
async def check_due_payments(context):
    logger.info("Checking due payments for notifications")
    today = datetime.now(ZoneInfo("Asia/Tashkent")).date()
    due_payments = await sync_to_async(list)(
        Pul_berish.objects.filter(
            Q(sana__lte=today) & Q(notification_sent=False) & Q(status="keyinroq_tolanadi")
        ).select_related('taminotchi', 'pul_olingan')
    )
    if not due_payments:
        logger.info("No due payments found")
        return
    bot = Bot(token="8002360442:AAHrdMm2lWTfKXRWkoeSmYR97sgNMAJtuM8")
    users = await sync_to_async(list)(CustomUser.objects.filter(can_add_new_users=True))
    for payment in due_payments:
        try:
            supplier = payment.taminotchi
            amount = payment.summa
            due_date = payment.sana
            related_debt = payment.pul_olingan
            debt_status = related_debt.status if related_debt else "N/A"
            debt_amount = related_debt.umumiy_miqdor if related_debt else 0
            debt_paid = related_debt.tolangan if related_debt else 0
            remaining_debt = debt_amount - debt_paid
            message = (
                f"⚠️ To‘lov eslatmasi:\n"
                f"Taminotchi: {supplier.taminotchi_ismi}\n"
                f"Summa: {amount} so‘m\n"
                f"To‘lov sanasi: {due_date}\n"
                f"Qarz holati: {remaining_debt} so‘m qoldi (Status: {debt_status})\n"
                f"Iltimos, to‘lovni tekshiring!"
            )
            for user in users:
                if user.telegram_chat_id:
                    try:
                        await bot.send_message(chat_id=user.telegram_chat_id, text=message)
                        logger.info(f"Notification sent to user {user.id} for payment {payment.id}")
                    except Exception as e:
                        logger.error(f"Failed to send notification to user {user.id}: {e}")
            payment.notification_sent = True
            await sync_to_async(payment.save)()
        except Exception as e:
            logger.error(f"Error processing payment {payment.id}: {e}")
    logger.info(f"Processed {len(due_payments)} due payments")

# Start scheduler
async def on_bot_start():
    logger.info("Bot starting, initializing scheduler")
    scheduler.add_job(check_due_payments, 'interval', days=1, start_date=datetime.now())
    scheduler.start()
# Utility function to check login status
async def check_login(context, update, reply_text=True):
    if 'user_id' in context.user_data:
        return True
    if reply_text:
        if update.callback_query:
            await update.callback_query.message.reply_text("Avval login qiling. /start buyrug‘ini bosing.")
        else:
            await update.message.reply_text("Avval login qiling. /start buyrug‘ini bosing.")
    return False

# Navigation keyboard
def nav_keyboard(extra=None):
    keyboard = []
    if extra:
        if isinstance(extra[0], list):
            keyboard.extend(extra)
        else:
            keyboard.append(extra)
    keyboard.append([
        InlineKeyboardButton("Ortga", callback_data="btn_back"),
        InlineKeyboardButton("Menu", callback_data="btn_menu"),
    ])
    return InlineKeyboardMarkup(keyboard)

# --- LOGIN ---
async def start(update, context):
    if await check_login(context, update, reply_text=False):
        await show_menu(update, context)
        return ConversationHandler.END
    await update.message.reply_text("Sizga berilgan login kodni kiriting: ")
    context.user_data['current_state'] = ASK_LOGIN
    return ASK_LOGIN

async def login_handler(update, context):
    login_code = update.message.text.strip()
    try:
        user = await sync_to_async(CustomUser.objects.get)(telegram_bot_login=login_code)
        context.user_data['user_id'] = user.id
        user.telegram_chat_id = str(update.message.chat_id)  # NEW
        await sync_to_async(user.save)()  # NEW
        context.user_data['current_state'] = None
        await update.message.reply_text("Salom, tizimga muvaffaqiyatli kirdingiz!")
        await show_menu(update, context)
        return ConversationHandler.END
    except CustomUser.DoesNotExist:
        await update.message.reply_text("Login kod noto'g'ri. Qayta urinib ko'ring:")
        return ASK_LOGIN
# --- LOGOUT ---
async def logout(update, context):
    context.user_data.clear()
    await update.message.reply_text("Tizimdan chiqdingiz. Qayta kirish uchun /start buyrug‘ini bosing.")
    return ConversationHandler.END
async def show_menu(update, context):
    if not await check_login(context, update):
        return ConversationHandler.END
    user_id = context.user_data.get('user_id')
    context.user_data.clear()
    context.user_data['user_id'] = user_id
    user = await sync_to_async(CustomUser.objects.get)(id=user_id)
    keyboard = [
        [InlineKeyboardButton("Taminotchilar", callback_data="btn_taminotchilar")],
        [
            InlineKeyboardButton("Hisobot", callback_data="btn_hisobot"),
            InlineKeyboardButton("Taminotchi qo'shish", callback_data="btn_taminotchi")
        ],
        [InlineKeyboardButton("Harajatlar", callback_data="btn_harajatlar")],
        [InlineKeyboardButton("Rejalashtirilgan to‘lovlar", callback_data="btn_rejala_tolovlar")], # O‘zgardi!
        [InlineKeyboardButton("Xabarlar", callback_data="btn_xabarlar")]
    ]
    if user.can_add_new_users:
        keyboard.insert(1, [InlineKeyboardButton("Ishchilar", callback_data="btn_ishchilar")])
    markup = InlineKeyboardMarkup(keyboard)
    text = "Menyudan tanlang:"
    if update.message:
        await update.message.reply_text(text, reply_markup=markup)
    else:
        await update.callback_query.message.reply_text(text, reply_markup=markup)
    return ConversationHandler.END


## ISHCHILAR

# NEW: ISHCHILAR Handlers
async def show_ishchilar(update, context):
    if not await check_login(context, update):
        return ConversationHandler.END
    user = await sync_to_async(CustomUser.objects.get)(id=context.user_data.get('user_id'))
    if not user.can_add_new_users:
        await (update.callback_query.message if update.callback_query else update.message).reply_text(
            "Sizda ishchilarni ko‘rish huquqi yo‘q.", reply_markup=nav_keyboard()
        )
        return ConversationHandler.END
    page = context.user_data.get('ishchi_page', 1)
    # Use filtered users_list if available, otherwise fetch all users
    users = context.user_data.get('users_list', None)
    if users is None:
        users = await sync_to_async(list)(CustomUser.objects.all().order_by('first_name', 'last_name'))
        context.user_data['users_list'] = users
    if not users:
        await (update.callback_query.message if update.callback_query else update.message).reply_text(
            "Hech qanday ishchi topilmadi.", reply_markup=nav_keyboard()
        )
        context.user_data['current_state'] = CHOOSE_ISHCHI
        return CHOOSE_ISHCHI
    total = len(users)
    start = (page - 1) * PAGE_SIZE
    end = start + PAGE_SIZE
    page_users = users[start:end]
    text = f"Ishchilar ro‘yxati ({start+1}-{min(end, total)}/Jami: {total}):\n"
    for idx, user in enumerate(page_users, start=1):
        text += f"{start+idx}. {user.first_name} {user.last_name} (Tel: {user.telefon_number or '-'})\n"
    row = []
    if start > 0:
        row.append(InlineKeyboardButton("⬅️ Oldingi", callback_data="prev_ishchilar"))
    if end < total:
        row.append(InlineKeyboardButton("➡️ Keyingi", callback_data="next_ishchilar"))
    keyboard = []
    if row:
        keyboard.append(row)
    keyboard.append([
        InlineKeyboardButton("Qidirish", callback_data="find_ishchi"),
        InlineKeyboardButton("Qo‘shish", callback_data="add_ishchi"),
    ])
    markup = nav_keyboard(extra=keyboard)
    if update.callback_query:
        await update.callback_query.edit_message_text(text, reply_markup=markup)
    else:
        await update.message.reply_text(text, reply_markup=markup)
    context.user_data['current_state'] = CHOOSE_ISHCHI
    return CHOOSE_ISHCHI

async def ishchilar_callback(update, context):
    query = update.callback_query
    await query.answer()
    if not await check_login(context, update):
        return ConversationHandler.END
    user = await sync_to_async(CustomUser.objects.get)(id=context.user_data.get('user_id'))
    if not user.can_add_new_users:
        await query.message.reply_text("Sizda ishchilarni ko‘rish huquqi yo‘q.", reply_markup=nav_keyboard())
        return ConversationHandler.END
    data = query.data
    if data == "prev_ishchilar":
        context.user_data['ishchi_page'] = max(context.user_data.get('ishchi_page', 1) - 1, 1)
        return await show_ishchilar(update, context)
    elif data == "next_ishchilar":
        context.user_data['ishchi_page'] = context.user_data.get('ishchi_page', 1) + 1
        return await show_ishchilar(update, context)
    elif data == "find_ishchi":
        await query.message.reply_text("Ism yoki familiya bo‘yicha qidirish uchun matn kiriting:", reply_markup=nav_keyboard())
        context.user_data['current_state'] = FIND_ISHCHI
        return FIND_ISHCHI
    elif data == "add_ishchi":
        await query.message.reply_text("Yangi ishchi ismini kiriting:", reply_markup=nav_keyboard())
        context.user_data['current_state'] = ISHCHI_NAME
        return ISHCHI_NAME

async def ishchi_name(update, context):
    context.user_data['new_user_first_name'] = update.message.text.strip()
    await update.message.reply_text("Familiyasini kiriting:", reply_markup=nav_keyboard())
    context.user_data['current_state'] = ISHCHI_SURNAME
    return ISHCHI_SURNAME

async def ishchi_surname(update, context):
    context.user_data['new_user_last_name'] = update.message.text.strip()
    await update.message.reply_text("Telefon raqamini kiriting (masalan, +998991234567):", reply_markup=nav_keyboard())
    context.user_data['current_state'] = ISHCHI_PHONE
    return ISHCHI_PHONE

async def ishchi_phone(update, context):
    phone = update.message.text.strip()
    if not re.match(r'^\+998\d{9}$', phone):
        await update.message.reply_text("Noto‘g‘ri telefon raqami! Format: +998991234567", reply_markup=nav_keyboard())
        return ISHCHI_PHONE
    context.user_data['new_user_phone'] = phone
    keyboard = [
        [InlineKeyboardButton("Ha", callback_data="perm_yes"), InlineKeyboardButton("Yo‘q", callback_data="perm_no")]
    ]
    markup = nav_keyboard(extra=keyboard)
    await update.message.reply_text("Harajat qo‘shish ruxsatini berasizmi?", reply_markup=markup)
    context.user_data['current_state'] = ISHCHI_PERMISSIONS
    context.user_data['perm_step'] = 'can_add_expanse'
    return ISHCHI_PERMISSIONS

# NEW: Ishchilar qo‘shimcha handler’lari
import uuid  # Random login uchun

async def ishchi_permissions(update, context):
    query = update.callback_query
    await query.answer()
    perm_step = context.user_data.get('perm_step')
    value = query.data == "perm_yes"
    context.user_data.setdefault('new_user_permissions', {})

    # Save the current permission value if a valid step is provided
    if perm_step:
        context.user_data['new_user_permissions'][perm_step] = value

    steps = [
        ('can_add_expanse_to_others', "Boshqalarga harajat qo‘shish ruxsatini berasizmi?"),
        ('can_add_new_users', "Yangi foydalanuvchi qo‘shish ruxsatini berasizmi?"),
        ('can_change_expanse', "Harajatlarni o‘zgartirish ruxsatini berasizmi?")
    ]

    # Find the index of the current perm_step
    current_index = -1
    if perm_step:
        for i, (step, _) in enumerate(steps):
            if step == perm_step:
                current_index = i
                break

    # Move to the next step
    next_index = current_index + 1

    # If there is a next step, ask the corresponding question
    if next_index < len(steps):
        next_step, question = steps[next_index]
        keyboard = [
            [InlineKeyboardButton("Ha", callback_data="perm_yes"), InlineKeyboardButton("Yo‘q", callback_data="perm_no")]
        ]
        markup = nav_keyboard(extra=keyboard)
        await query.message.reply_text(question, reply_markup=markup)
        context.user_data['perm_step'] = next_step
        context.user_data['current_state'] = ISHCHI_PERMISSIONS
        return ISHCHI_PERMISSIONS

    # All permissions have been asked, show confirmation
    user_data = context.user_data
    text = (
        f"Ishchi ma’lumotlari:\n"
        f"Ism: {user_data['new_user_first_name']}\n"
        f"Familiya: {user_data['new_user_last_name']}\n"
        f"Telefon: {user_data['new_user_phone']}\n"
        f"Ruxsatlar:\n"
        f"- Harajat qo‘shish: {'Ha' if user_data['new_user_permissions'].get('can_add_expanse') else 'Yo‘q'}\n"
        f"- Boshqalarga harajat: {'Ha' if user_data['new_user_permissions'].get('can_add_expanse_to_others') else 'Yo‘q'}\n"
        f"- Yangi foydalanuvchi: {'Ha' if user_data['new_user_permissions'].get('can_add_new_users') else 'Yo‘q'}\n"
        f"- Harajat o‘zgartirish: {'Ha' if user_data['new_user_permissions'].get('can_change_expanse') else 'Yo‘q'}\n"
        f"Ma’lumotlarni tasdiqlaysizmi?"
    )
    keyboard = [
        [InlineKeyboardButton("✅ Ha", callback_data="confirm_yes"), InlineKeyboardButton("❌ Yo‘q", callback_data="confirm_no")]
    ]
    markup = nav_keyboard(extra=keyboard)
    await query.message.reply_text(text, reply_markup=markup)
    context.user_data['current_state'] = ISHCHI_CONFIRM
    return ISHCHI_CONFIRM
async def ishchi_confirm(update, context):
    query = update.callback_query
    await query.answer()
    if query.data == "confirm_yes":
        user_data = context.user_data
        try:
            # Unikal loginni to'g'ri yaratish
            login = str(uuid.uuid4())[:5]
            # Yangi userni yaratishdan oldin, login band emasligiga ishonch hosil qilamiz
            from asgiref.sync import sync_to_async
            is_exists = await sync_to_async(CustomUser.objects.filter(telegram_bot_login=login).exists)()
            while is_exists:
                login = str(uuid.uuid4())[:5]
                is_exists = await sync_to_async(CustomUser.objects.filter(telegram_bot_login=login).exists)()
            
            # (Agar username ham unique bo'lsa, uni ham shunaqa tekshiring)

            new_user = await sync_to_async(CustomUser.objects.create)(
                first_name=user_data['new_user_first_name'],
                last_name=user_data['new_user_last_name'],
                telefon_number=user_data['new_user_phone'],
                username = login,
                telegram_bot_login=login,
                can_add_expanse=user_data['new_user_permissions'].get('can_add_expanse', False),
                can_add_expanse_to_others=user_data['new_user_permissions'].get('can_add_expanse_to_others', False),
                can_add_new_users=user_data['new_user_permissions'].get('can_add_new_users', False),
                can_change_expanse=user_data['new_user_permissions'].get('can_change_expanse', False)
            )
            await query.message.reply_text(
                f"Ishchi {new_user.first_name} qo‘shildi! Login: {new_user.telegram_bot_login}",
                reply_markup=nav_keyboard()
            )
            context.user_data['current_state'] = None
            await show_menu(update, context)
            return ConversationHandler.END
        except Exception as e:
            logger.error(f"Error creating user: {e}")
            await query.message.reply_text("Xatolik yuz berdi. Qayta urinib ko‘ring.", reply_markup=nav_keyboard())
            return ISHCHI_CONFIRM
    elif query.data == "confirm_no":
        context.user_data['current_state'] = None
        await show_menu(update, context)
        return ConversationHandler.END

async def find_ishchi(update, context):
    if not await check_login(context, update):
        return ConversationHandler.END
    search_text = update.message.text.strip().lower()
    users = await sync_to_async(list)(
        CustomUser.objects.filter(
            models.Q(first_name__icontains=search_text) | models.Q(last_name__icontains=search_text)
        ).order_by('first_name', 'last_name')
    )
    context.user_data['users_list'] = users
    if not users:
        await update.message.reply_text(
            "Qidiruvga mos ishchi topilmadi.", reply_markup=nav_keyboard()
        )
        await update.message.reply_text("Harajat miqdorini kiriting:", reply_message=nav_keyboard())
        context.user_data['current_state'] = CHOOSE_ISHCHI
        return await show_ishchilar(update, context)
    text = f"Qidiruv natijalari ({len(users)} ta):\n"
    for idx, user in enumerate(users, start=1):
        text += f"{idx}. {user.first_name} {user.last_name} (Tel: {user.telefon_number or '-'})\n"
    keyboard = []
    if len(users) > PAGE_SIZE:
        keyboard.append([
            InlineKeyboardButton("⬅️ Oldingi", callback_data="prev_ishchilar"),
            InlineKeyboardButton("➡️", callback_data="next_ishchilar")
        ])
    keyboard.append(
        [InlineKeyboardButton("Qidirish", callback_data="find_ishchi"),
        InlineKeyboardButton("Qo‘shish", callback_data="add_ishchi"),
    ])
    markup = nav_keyboard(extra=keyboard)
    await update.message.reply_text(text, reply_markup=markup)
    context.user_data['current_state'] = CHOOSE_ISHCHI
    context.user_data['ishchi_page'] = 1
    return await show_ishchilar(update, context)
async def choose_ishchi(update, context):
    text = update.message.text.strip()
    users = context.user_data.get('users_list', [])
    if not users:
        users = await sync_to_async(list)(CustomUser.objects.all().order_by('first_name', 'last_name'))
        context.user_data['users_list'] = users
    if not users:
        await update.message.reply_text("Hech qanday ishchi topilmadi.", reply_markup=nav_keyboard())
        context.user_data['current_state'] = CHOOSE_ISHCHI
        return CHOOSE_ISHCHI
    if text.isdigit():
        idx = int(text) - 1
        if 0 <= idx < len(users):
            context.user_data['selected_user'] = users[idx]
            context.user_data['current_state'] = ISHCHI_ACTION
            return await show_ishchi_info(update, context)
        else:
            await update.message.reply_text("Noto‘g‘ri raqam. Qaytadan tanlang:", reply_markup=nav_keyboard())
            return CHOOSE_ISHCHI
    await update.message.reply_text("Faqat ro‘yxatdan raqam kiriting!", reply_markup=nav_keyboard())
    return CHOOSE_ISHCHI

async def show_ishchi_info(update, context):
    user = context.user_data.get('selected_user')
    if not user:
        await (update.callback_query.message if update.callback_query else update.message).reply_text(
            "Ishchi tanlanmadi.", reply_markup=nav_keyboard()
        )
        context.user_data['current_state'] = None
        await show_menu(update, context)
        return ConversationHandler.END
    text = (
        f"Ishchi: {user.first_name} {user.last_name}\n"
        f"Telefon: {user.telefon_number or '-'}\n"
        f"Login: {user.telegram_bot_login}\n"
        f"Ruxsatlar:\n"
        f"- Harajat qo‘shish: {'Ha' if user.can_add_expanse else 'Yo‘q'}\n"
        f"- Boshqalarga harajat: {'Ha' if user.can_add_expanse_to_others else 'Yo‘q'}\n"
        f"- Yangi foydalanuvchi: {'Ha' if user.can_add_new_users else 'Yo‘q'}\n"
        f"- Harajat o‘zgartirish: {'Ha' if user.can_change_expanse else 'Yo‘q'}\n"
    )
    keyboard = [
        [
            InlineKeyboardButton("O‘chirish", callback_data="delete_ishchi"),
            InlineKeyboardButton("Tahrirlash", callback_data="edit_ishchi")
        ]
    ]
    markup = nav_keyboard(extra=keyboard)
    if update.callback_query:
        await update.callback_query.message.reply_text(text, reply_markup=markup)
    else:
        await update.message.reply_text(text, reply_markup=markup)
    context.user_data['current_state'] = ISHCHI_ACTION
    return ISHCHI_ACTION

async def ishchi_action_handler(update, context):
    query = update.callback_query
    await query.answer()
    if query.data == "delete_ishchi":
        user = context.user_data.get('selected_user')
        text = f"{user.first_name} {user.last_name} ni o‘chirishni tasdiqlaysizmi?"
        keyboard = [
            [InlineKeyboardButton("✅ Ha", callback_data="confirm_delete"), InlineKeyboardButton("❌ Yo‘q", callback_data="cancel_delete")]
        ]
        markup = nav_keyboard(extra=keyboard)
        await query.message.reply_text(text, reply_markup=markup)
        context.user_data['current_state'] = ISHCHI_DELETE_CONFIRM
        return ISHCHI_DELETE_CONFIRM
    elif query.data == "edit_ishchi":
        await query.message.reply_text("Yangi ismini kiriting:", reply_markup=nav_keyboard())
        context.user_data['current_state'] = ISHCHI_EDIT_NAME
        return ISHCHI_EDIT_NAME

async def ishchi_delete_confirm_handler(update, context):
    query = update.callback_query
    await query.answer()
    if query.data == "confirm_delete":
        user = context.user_data.get('selected_user')
        try:
            await sync_to_async(user.delete)()
            await query.message.reply_text(f"{user.first_name} {user.last_name} o‘chirildi.", reply_markup=nav_keyboard())
            context.user_data['current_state'] = None
            return await show_ishchilar(update, context)
        except Exception as e:
            logger.error(f"Error deleting user: {e}")
            await query.message.reply_text("Xatolik yuz berdi. Qayta urinib ko‘ring.", reply_markup=nav_keyboard())
            return ISHCHI_DELETE_CONFIRM
    elif query.data == "cancel_delete":
        return await show_ishchi_info(update, context)

async def ishchi_edit_name(update, context):
    context.user_data['edit_user_first_name'] = update.message.text.strip()
    await update.message.reply_text("Yangi familiyasini kiriting:", reply_markup=nav_keyboard())
    context.user_data['current_state'] = ISHCHI_EDIT_SURNAME
    return ISHCHI_EDIT_SURNAME

async def ishchi_edit_surname(update, context):
    context.user_data['edit_user_last_name'] = update.message.text.strip()
    await update.message.reply_text("Yangi telefon raqamini kiriting (masalan, +998991234567):", reply_markup=nav_keyboard())
    context.user_data['current_state'] = ISHCHI_EDIT_PHONE
    return ISHCHI_EDIT_PHONE

async def ishchi_edit_phone(update, context):
    phone = update.message.text.strip()
    if not re.match(r'^\+998\d{9}$', phone):
        await update.message.reply_text("Noto‘g‘ri telefon raqami! Format: +998991234567", reply_markup=nav_keyboard())
        return ISHCHI_EDIT_PHONE
    context.user_data['edit_user_phone'] = phone
    keyboard = [
        [InlineKeyboardButton("Ha", callback_data="perm_yes"), InlineKeyboardButton("Yo‘q", callback_data="perm_no")]
    ]
    markup = nav_keyboard(extra=keyboard)
    await update.message.reply_text("Harajat qo‘shish ruxsatini berasizmi?", reply_markup=markup)
    context.user_data['current_state'] = ISHCHI_EDIT_PERMISSIONS
    context.user_data['perm_step'] = 'can_add_expanse'
    return ISHCHI_EDIT_PERMISSIONS
async def ishchi_edit_permissions(update, context):
    query = update.callback_query
    await query.answer()
    perm_step = context.user_data.get('perm_step')
    value = query.data == "perm_yes"
    context.user_data.setdefault('edit_user_permissions', {})
    
    # Save the current permission value if perm_step exists
    if perm_step:
        context.user_data['edit_user_permissions'][perm_step] = value

    steps = [
        ('can_add_expanse', "Harajat qo‘shish ruxsatini berasizmi?"),
        ('can_add_expanse_to_others', "Boshqalarga harajat qo‘shish ruxsatini berasizmi?"),
        ('can_add_new_users', "Yangi foydalanuvchi qo‘shish ruxsatini berasizmi?"),
        ('can_change_expanse', "Harajatlarni o‘zgartirish ruxsatini berasizmi?")
    ]

    # Find the current step index
    current_index = -1
    if perm_step:
        for i, (step, _) in enumerate(steps):
            if step == perm_step:
                current_index = i
                break

    # Move to the next step
    next_index = current_index + 1

    # If there is a next step, ask the corresponding question
    if next_index < len(steps):
        next_step, question = steps[next_index]
        keyboard = [
            [InlineKeyboardButton("Ha", callback_data="perm_yes"), InlineKeyboardButton("Yo‘q", callback_data="perm_no")]
        ]
        markup = nav_keyboard(extra=keyboard)
        await query.message.reply_text(question, reply_markup=markup)
        context.user_data['perm_step'] = next_step
        context.user_data['current_state'] = ISHCHI_EDIT_PERMISSIONS
        return ISHCHI_EDIT_PERMISSIONS

    # All permissions have been asked, show confirmation
    user_data = context.user_data
    user = user_data.get('selected_user')
    text = (
        f"Tahrirlangan ma’lumotlar:\n"
        f"Ism: {user_data['edit_user_first_name']}\n"
        f"Familiya: {user_data['edit_user_last_name']}\n"
        f"Telefon: {user_data['edit_user_phone']}\n"
        f"Ruxsatlar:\n"
        f"- Harajat qo‘shish: {'Ha' if user_data['edit_user_permissions'].get('can_add_expanse') else 'Yo‘q'}\n"
        f"- Boshqalarga harajat: {'Ha' if user_data['edit_user_permissions'].get('can_add_expanse_to_others') else 'Yo‘q'}\n"
        f"- Yangi foydalanuvchi: {'Ha' if user_data['edit_user_permissions'].get('can_add_new_users') else 'Yo‘q'}\n"
        f"- Harajat o‘zgartirish: {'Ha' if user_data['edit_user_permissions'].get('can_change_expanse') else 'Yo‘q'}\n"
        f"Ma’lumotlarni tasdiqlaysizmi?"
    )
    keyboard = [
        [InlineKeyboardButton("✅ Ha", callback_data="confirm_yes"), InlineKeyboardButton("❌ Yo‘q", callback_data="confirm_no")]
    ]
    markup = nav_keyboard(extra=keyboard)
    await query.message.reply_text(text, reply_markup=markup)
    context.user_data['current_state'] = ISHCHI_EDIT_CONFIRM
    return ISHCHI_EDIT_CONFIRM

async def ishchi_edit_confirm(update, context):
    query = update.callback_query
    await query.answer()
    if query.data == "confirm_yes":
        user = context.user_data.get('selected_user')
        user_data = context.user_data
        try:
            user.first_name = user_data['edit_user_first_name']
            user.last_name = user_data['edit_user_last_name']
            user.telefon_number = user_data['edit_user_phone']
            user.can_add_expanse = user_data['edit_user_permissions'].get('can_add_expanse', False)
            user.can_add_expanse_to_others = user_data['edit_user_permissions'].get('can_add_expanse_to_others', False)
            user.can_add_new_users = user_data['edit_user_permissions'].get('can_add_new_users', False)
            user.can_change_expanse = user_data['edit_user_permissions'].get('can_change_expanse', False)
            await sync_to_async(user.save)()
            await query.message.reply_text(f"{user.first_name} {user.last_name} ma’lumotlari yangilandi.", reply_markup=nav_keyboard())
            context.user_data['current_state'] = None
            return await show_ishchilar(update, context)
        except Exception as e:
            logger.error(f"Error updating user: {e}")
            await query.message.reply_text("Xatolik yuz berdi. Qayta urinib ko‘ring.", reply_markup=nav_keyboard())
            return ISHCHI_EDIT_CONFIRM
    elif query.data == "confirm_no":
        return await show_ishchi_info(update, context)
# Qolgan Ishchilar handler’lari (qisqa ko‘rinish):
# - ishchi_permissions: can_add_expanse, can_add_expanse_to_others, can_add_new_users, can_change_expanse uchun ketma-ket so‘raydi.
# - ishchi_confirm: Ma’lumotlarni tasdiqlash.
# - choose_ishchi, find_ishchi, show_ishchi_info, ishchi_action, ishchi_delete_confirm, tahrirlash funksiyalari.
async def harajat_type_handler(update, context):
    query = update.callback_query
    await query.answer()
    if not await check_login(context, update):
        return await show_menu(update, context)
    user = await sync_to_async(CustomUser.objects.get())(id=context.user_data.get('user_id'))
    data = query.data
    if data == "harajat_ozimga":
        if not user.can_add_expanse:
            await query.message.reply_text("Sizda harajat qo‘shish huquqi yo‘q.", reply_markup=nav_keyboard())
            return await show_menu(update, context)
        context.user_data['harajat_type'] = 'ozimga'
        context.user_data['harajat_ishchi'] = user
        await query.message.reply_text("Harajat miqdorini kiriting:", reply_markup=nav_keyboard())
        context.user_data['current_state'] = HARAJAT_AMOUNT
        logger.info(f"Transition to HARAJAT_AMOUNT: {HARAJAT_AMOUNT}")
        return HARAJAT_AMOUNT
    elif data == "harajat_ishxonaga":
        if not user.can_add_expanse:
            await query.message.reply_text("Sizda harajat qo‘shish huquqi yo‘q.", reply_markup=nav_keyboard())
            return await show_menu(update, context)
        context.user_data['harajat_type'] = 'ishxonaga'
        context.user_data['harajat_ishchi'] = None
        await query.message.reply_text("Harajat miqdorini kiriting:", reply_markup=nav_keyboard())
        context.user_data['current_state'] = HARAJAT_AMOUNT
        logger.info(f"Transition to HARAJAT_AMOUNT: {HARAJAT_AMOUNT}")
        return HARAJAT_AMOUNT
    elif data == "harajat_ishchilarga":
        if not user.can_add_expanse_to_others:
            await query.message.reply_text("Sizda boshqalarga harajat qo‘shish huquqi yo‘q.", reply_markup=nav_keyboard())
            return await show_menu(update, context)
        context.user_data['harajat_type'] = 'ishchilarga'
        context.user_data['ishchi_page'] = 1
        logger.info("Calling show_ishchilar_for_harajat")
        return await show_ishchilar_for_harajat(update, context)
# NEW: HARAJATLAR Handlers
async def harajat_menu_handler(update, context):
    query = update.callback_query
    await query.answer()
    if not await check_login(context, update):
        return ConversationHandler.END
    user = await sync_to_async(CustomUser.objects.get)(id=context.user_data.get('user_id'))
    data = query.data
    if data in ["harajat_ozimga", "harajat_ishxonaga"]:
        if not user.can_add_expanse:
            await query.message.reply_text("Sizda harajat qo‘shish huquqi yo‘q.", reply_markup=nav_keyboard())
            return ConversationHandler.END
        context.user_data['harajat_type'] = 'ozimga' if data == "harajat_ozimga" else 'ishxonaga'
        context.user_data['harajat_ishchi'] = user if data == "harajat_ozimga" else None
        await query.message.reply_text("Harajat miqdorini kiriting:", reply_markup=nav_keyboard())
        context.user_data['current_state'] = HARAJAT_AMOUNT
        return HARAJAT_AMOUNT
    elif data == "harajat_ishchilarga":
        if not user.can_add_expanse_to_others:
            await query.message.reply_text("Sizda boshqalarga harajat qo‘shish huquqi yo‘q.", reply_markup=nav_keyboard())
            return ConversationHandler.END
        context.user_data['harajat_type'] = 'ishchilarga'
        context.user_data['ishchi_page'] = 1
        return await show_ishchilar_for_harajat(update, context)
    elif data == "harajat_list":
        context.user_data['harajat_page'] = 1
        return await show_harajat_list(update, context)
    elif data == "btn_back":
        context.user_data['current_state'] = None
        await show_menu(update, context)
        return ConversationHandler.END

async def show_ishchilar_for_harajat(update, context):
    page = context.user_data.get('ishchi_page', 1)
    users = await sync_to_async(list)(CustomUser.objects.all().order_by('first_name', 'last_name'))
    if not users:
        await (update.callback_query.message if update.callback_query else update.message).reply_text(
            "Hech qanday ishchi topilmadi.", reply_markup=nav_keyboard()
        )
        context.user_data['current_state'] = HARAJAT_MENU
        return HARAJAT_MENU
    total = len(users)
    start = (page - 1) * PAGE_SIZE
    end = start + PAGE_SIZE
    page_users = users[start:end]
    context.user_data['users_list'] = users
    text = f"Ishchilar ro‘yxati ({start+1}-{min(end, total)}/Jami: {total}):\n"
    for idx, user in enumerate(page_users, start=1):
        text += f"{start+idx}. {user.first_name} {user.last_name} (Tel: {user.telefon_number or '-'})\n"
    row = []
    if start > 0:
        row.append(InlineKeyboardButton("⬅️ Oldingi", callback_data="prev_ishchilar_harajat"))
    if end < total:
        row.append(InlineKeyboardButton("➡️ Keyingi", callback_data="next_ishchilar_harajat"))
    keyboard = []
    if row: keyboard.append(row)
    markup = nav_keyboard(extra=keyboard)
    if update.callback_query:
        await update.callback_query.edit_message_text(text, reply_markup=markup)
    else:
        await update.message.reply_text(text, reply_markup=markup)
    context.user_data['current_state'] = CHOOSE_ISHCHI_EXPENSE
    return CHOOSE_ISHCHI_EXPENSE

async def harajat_amount(update, context):
    logger.info(f"harajat_amount called, returning state: {HARAJAT_AMOUNT}")
    try:
        amount = float(update.message.text.replace(",", "."))
        if amount <= 0:
            await update.message.reply_text("Miqdor musbat bo‘lishi kerak. Qaytadan kiriting:", reply_markup=nav_keyboard())
            return HARAJAT_AMOUNT
        context.user_data['harajat_amount'] = amount
        await update.message.reply_text("Harajat sanasini kiriting (DD-MM-YYYY):", reply_markup=nav_keyboard())
        context.user_data['current_state'] = HARAJAT_DATE
        logger.info(f"Moving to HARAJAT_DATE: {HARAJAT_DATE}")
        return HARAJAT_DATE
    except ValueError:
        await update.message.reply_text("Noto‘g‘ri miqdor. Qaytadan kiriting:", reply_markup=nav_keyboard())
        return HARAJAT_AMOUNT
# NEW: Harajatlar qo‘shimcha handler’lari
async def choose_ishchi_expense(update, context):
    text = update.message.text.strip()
    users = context.user_data.get('users_list', [])
    if not users:
        users = await sync_to_async(list)(CustomUser.objects.all().order_by('first_name', 'last_name'))
        context.user_data['users_list'] = users
    if not users:
        await update.message.reply_text("Hech qanday ishchi topilmadi.", reply_markup=nav_keyboard())
        context.user_data['current_state'] = HARAJAT_MENU
        return HARAJAT_MENU
    if text.isdigit():
        idx = int(text) - 1
        if 0 <= idx < len(users):
            context.user_data['harajat_ishchi'] = users[idx]
            await update.message.reply_text("Harajat miqdorini kiriting:", reply_markup=nav_keyboard())
            context.user_data['current_state'] = HARAJAT_AMOUNT
            return HARAJAT_AMOUNT
        else:
            await update.message.reply_text("Noto‘g‘ri raqam. Qaytadan tanlang:", reply_markup=nav_keyboard())
            return CHOOSE_ISHCHI_EXPENSE
    await update.message.reply_text("Faqat ro‘yxatdan raqam kiriting!", reply_markup=nav_keyboard())
    return CHOOSE_ISHCHI_EXPENSE

async def harajat_callback(update, context):
    query = update.callback_query
    await query.answer()
    if query.data == "prev_ishchilar_harajat":
        context.user_data['ishchi_page'] = max(context.user_data.get('ishchi_page', 1) - 1, 1)
        return await show_ishchilar_for_harajat(update, context)
    elif query.data == "next_ishchilar_harajat":
        context.user_data['ishchi_page'] = context.user_data.get('ishchi_page', 1) + 1
        return await show_ishchilar_for_harajat(update, context)

async def harajat_date(update, context):
    date_text = update.message.text.strip()
    try:
        datetime.strptime(date_text, "%d-%m-%Y")
        context.user_data['harajat_date'] = date_text
        await update.message.reply_text("Sababni kiriting (ixtiyoriy, '-' deb yuboring):", reply_markup=nav_keyboard())
        context.user_data['current_state'] = HARAJAT_REASON
        return HARAJAT_REASON
    except ValueError:
        await update.message.reply_text("Sana formati noto‘g‘ri! To‘g‘ri format: DD-MM-YYYY.", reply_markup=nav_keyboard())
        return HARAJAT_DATE

async def harajat_reason(update, context):
    reason = update.message.text.strip()
    context.user_data['harajat_reason'] = "" if reason == "-" else reason
    user_data = context.user_data
    ishchi = user_data.get('harajat_ishchi')
    text = (
        f"Harajat ma’lumotlari:\n"
        f"Kim uchun: {ishchi.first_name + ' ' + ishchi.last_name if ishchi else 'Ishxona'}\n"
        f"Miqdor: {user_data['harajat_amount']}\n"
        f"Sana: {user_data['harajat_date']}\n"
        f"Sabab: {user_data['harajat_reason'] or '-'}\n"
        f"Ma’lumotlarni tasdiqlaysizmi?"
    )
    keyboard = [
        [InlineKeyboardButton("✅ Ha", callback_data="confirm_yes"), InlineKeyboardButton("❌ Yo‘q", callback_data="confirm_no")]
    ]
    markup = nav_keyboard(extra=keyboard)
    await update.message.reply_text(text, reply_markup=markup)
    context.user_data['current_state'] = HARAJAT_CONFIRMATION
    return HARAJAT_CONFIRMATION

async def harajat_confirmation(update, context):
    query = update.callback_query
    await query.answer()
    if query.data == "confirm_yes":
        user_data = context.user_data
        try:
            await sync_to_async(Harajatlar.objects.create)(
                ishchi=user_data.get('harajat_ishchi'),
                summa=user_data['harajat_amount'],
                sabab=user_data['harajat_reason'],
                sana=datetime.strptime(user_data['harajat_date'], "%d-%m-%Y").date()
            )
            await query.message.reply_text("Harajat qo‘shildi!", reply_markup=nav_keyboard())
            context.user_data['current_state'] = None
            await show_menu(update, context)
            return ConversationHandler.END
        except Exception as e:
            logger.error(f"Error creating expense: {e}")
            await query.message.reply_text("Xatolik yuz berdi. Qayta urinib ko‘ring.", reply_markup=nav_keyboard())
            return HARAJAT_CONFIRMATION
    elif query.data == "confirm_no":
        context.user_data['current_state'] = None
        await show_menu(update, context)
        return ConversationHandler.END
    
async def choose_harajat_by_number(update, context):
    logger.info(f"choose_harajat_by_number called with text: {update.message.text}")
    text = update.message.text.strip()
    harajatlar = context.user_data.get('harajatlar_list', [])
    if not harajatlar:
        harajatlar = await sync_to_async(list)(
            Harajatlar.objects.select_related('ishchi').all().order_by('-sana')
        )
        context.user_data['harajatlar_list'] = harajatlar
    if not harajatlar:
        await update.message.reply_text("Hech qanday harajat topilmadi.", reply_markup=nav_keyboard())
        context.user_data['current_state'] = HARAJAT_MENU
        return HARAJAT_MENU
    if text.isdigit():
        idx = int(text) - 1
        page = context.user_data.get('harajat_page', 1)
        start = (page - 1) * PAGE_SIZE
        global_idx = start + idx
        if 0 <= global_idx < len(harajatlar):
            harajat = harajatlar[global_idx]
            context.user_data['selected_harajat'] = harajat
            user = await sync_to_async(CustomUser.objects.get)(id=context.user_data.get('user_id'))
            keyboard = []
            if user.can_change_expanse:
                keyboard.append([
                    InlineKeyboardButton("O‘chirish", callback_data="delete_harajat"),
                    InlineKeyboardButton("Tahrirlash", callback_data="edit_harajat")
                ])
            markup = nav_keyboard(extra=keyboard)
            kim = harajat.ishchi.first_name + " " + harajat.ishchi.last_name if harajat.ishchi else "Ishxona"
            text = (
                f"Harajat:\n"
                f"Kim uchun: {kim}\n"
                f"Miqdor: {harajat.summa}\n"
                f"Sana: {harajat.sana}\n"
                f"Sabab: {harajat.sabab or '-'}\n"
            )
            await update.message.reply_text(text, reply_markup=markup)
            context.user_data['current_state'] = HARAJAT_ACTION
            logger.info(f"Selected harajat ID: {harajat.id}, transitioned to HARAJAT_ACTION")
            return HARAJAT_ACTION
        else:
            await update.message.reply_text("Noto‘g‘ri raqam. Qaytadan tanlang:", reply_markup=nav_keyboard())
            return HARAJAT_LIST
    await update.message.reply_text("Faqat ro‘yxatdan raqam kiriting!", reply_markup=nav_keyboard())
    return HARAJAT_LIST

async def show_harajat_list(update, context):
    logger.info("show_harajat_list called")
    page = context.user_data.get('harajat_page', 1)
    harajatlar = await sync_to_async(list)(
        Harajatlar.objects.select_related('ishchi').all().order_by('-sana')
    )
    if not harajatlar:
        await (update.callback_query.message if update.callback_query else update.message).reply_text(
            "Hech qanday harajat topilmadi.", reply_markup=nav_keyboard()
        )
        context.user_data['current_state'] = HARAJAT_MENU
        return HARAJAT_MENU
    total = len(harajatlar)
    start = (page - 1) * PAGE_SIZE
    end = start + PAGE_SIZE
    page_harajatlar = harajatlar[start:end]
    context.user_data['harajatlar_list'] = harajatlar
    text = f"Harajatlar ro‘yxati ({start+1}-{min(end, total)}/Jami: {total}):\n"
    for idx, harajat in enumerate(page_harajatlar, start=1):
        kim = await sync_to_async(lambda: harajat.ishchi.first_name + " " + harajat.ishchi.last_name if harajat.ishchi else "Ishxona")()
        text += f"{start+idx}. {kim} - {harajat.summa} ({harajat.sana})\n"
    text += "\nHarajatni tanlash uchun uning raqamini yuboring (masalan, 1):"  # NEW: Prompt for number input
    row = []
    if start > 0:
        row.append(InlineKeyboardButton("⬅️ Oldingi", callback_data="prev_harajat"))
    if end < total:
        row.append(InlineKeyboardButton("➡️ Keyingi", callback_data="next_harajat"))
    keyboard = []
    if row:
        keyboard.append(row)
    keyboard.append([InlineKeyboardButton("Excel’ga eksport", callback_data="harajat_excel")])
    markup = nav_keyboard(extra=keyboard)
    if update.callback_query:
        await update.callback_query.message.edit_text(text, reply_markup=markup)
    else:
        await update.message.reply_text(text, reply_markup=markup)
    context.user_data['current_state'] = HARAJAT_LIST
    logger.info("Rendering harajat list with Excel export button")
    return HARAJAT_LIST
async def harajat_list_callback(update, context):
    query = update.callback_query
    await query.answer()
    logger.info(f"harajat_list_callback called with data: {query.data}, current_state: {context.user_data.get('current_state')}")
    try:
        if query.data == "prev_harajat":
            context.user_data['harajat_page'] = max(context.user_data.get('harajat_page', 1) - 1, 1)
            logger.info(f"Navigating to previous page: {context.user_data['harajat_page']}")
            return await show_harajat_list(update, context)
        elif query.data == "next_harajat":
            context.user_data['harajat_page'] = context.user_data.get('harajat_page', 1) + 1
            logger.info(f"Navigating to next page: {context.user_data['harajat_page']}")
            return await show_harajat_list(update, context)
        elif query.data == "harajat_excel":
            logger.info("Excel export button clicked")
            keyboard = [
                [InlineKeyboardButton("Butun davr", callback_data="excel_full")],
                [InlineKeyboardButton("Vaqt oralig‘i bo‘yicha", callback_data="excel_interval")]
            ]
            markup = nav_keyboard(extra=keyboard)
            await query.message.reply_text("Eksport qaysi davr uchun kerak?", reply_markup=markup)
            context.user_data['current_state'] = HARAJAT_EXCEL_MENU
            logger.info(f"Transitioned to HARAJAT_EXCEL_MENU: {HARAJAT_EXCEL_MENU}")
            return HARAJAT_EXCEL_MENU
        elif query.data.startswith("harajat_action_"):
            harajat_id = int(query.data.split("_")[-1])
            logger.info(f"Selected harajat ID: {harajat_id}")
            harajat = await sync_to_async(Harajatlar.objects.select_related('ishchi').get)(id=harajat_id)
            context.user_data['selected_harajat'] = harajat
            user = await sync_to_async(CustomUser.objects.get)(id=context.user_data.get('user_id'))
            keyboard = []
            if user.can_change_expanse:
                keyboard.append([
                    InlineKeyboardButton("O‘chirish", callback_data="delete_harajat"),
                    InlineKeyboardButton("Tahrirlash", callback_data="edit_harajat")
                ])
            markup = nav_keyboard(extra=keyboard)
            kim = await sync_to_async(lambda: harajat.ishchi.first_name + " " + harajat.ishchi.last_name if harajat.ishchi else "Ishxona")()
            text = (
                f"Harajat:\n"
                f"Kim uchun: {kim}\n"
                f"Miqdor: {harajat.summa}\n"
                f"Sana: {harajat.sana}\n"
                f"Sabab: {harajat.sabab or '-'}\n"
            )
            await query.message.reply_text(text, reply_markup=markup)
            context.user_data['current_state'] = HARAJAT_ACTION
            logger.info(f"Transitioned to HARAJAT_ACTION: {HARAJAT_ACTION}")
            return HARAJAT_ACTION
        else:
            logger.warning(f"Unhandled callback data: {query.data}")
            await query.message.reply_text("Noto‘g‘ri buyruq, qaytadan urinib ko‘ring.", reply_markup=nav_keyboard())
            return HARAJAT_LIST
    except Exception as e:
        logger.error(f"Error in harajat_list_callback: {e}", exc_info=True)
        await query.message.reply_text("Xatolik yuz berdi, qayta urinib ko‘ring.", reply_markup=nav_keyboard())
        return HARAJAT_LIST

async def harajat_action_handler(update, context):
    query = update.callback_query
    await query.answer()
    user = await sync_to_async(CustomUser.objects.get)(id=context.user_data.get('user_id'))
    if not user.can_change_expanse:
        await query.message.reply_text("Sizda harajatlarni o‘zgartirish huquqi yo‘q.", reply_markup=nav_keyboard())
        return await show_harajat_list(update, context)
    if query.data == "delete_harajat":
        harajat = context.user_data.get('selected_harajat')
        kim = harajat.ishchi.first_name + " " + harajat.ishchi.last_name if harajat.ishchi else "Ishxona"
        text = f"{kim} uchun {harajat.summa} ({harajat.sana}) harajatni o‘chirishni tasdiqlaysizmi?"
        keyboard = [
            [InlineKeyboardButton("✅ Ha", callback_data="confirm_delete"), InlineKeyboardButton("❌ Yo‘q", callback_data="cancel_delete")]
        ]
        markup = nav_keyboard(extra=keyboard)
        await query.message.reply_text(text, reply_markup=markup)
        context.user_data['current_state'] = HARAJAT_DELETE_CONFIRM
        return HARAJAT_DELETE_CONFIRM
    elif query.data == "edit_harajat":
        await query.message.reply_text("Yangi miqdorni kiriting:", reply_markup=nav_keyboard())
        context.user_data['current_state'] = HARAJAT_EDIT_AMOUNT
        return HARAJAT_EDIT_AMOUNT

async def harajat_delete_confirm(update, context):
    query = update.callback_query
    await query.answer()
    if query.data == "confirm_delete":
        harajat = context.user_data.get('selected_harajat')
        try:
            await sync_to_async(harajat.delete)()
            await query.message.reply_text("Harajat o‘chirildi.", reply_markup=nav_keyboard())
            context.user_data['current_state'] = None
            return await show_harajat_list(update, context)
        except Exception as e:
            logger.error(f"Error deleting expense: {e}")
            await query.message.reply_text("Xatolik yuz berdi. Qayta urinib ko‘ring.", reply_markup=nav_keyboard())
            return HARAJAT_DELETE_CONFIRM
    elif query.data == "cancel_delete":
        context.user_data['current_state'] = HARAJAT_ACTION
        return await harajat_action_handler(update, context)

async def harajat_edit_amount(update, context):
    try:
        amount = float(update.message.text.replace(",", "."))
        if amount <= 0:
            await update.message.reply_text("Miqdor musbat bo‘lishi kerak. Qaytadan kiriting:", reply_markup=nav_keyboard())
            return HARAJAT_EDIT_AMOUNT
        context.user_data['edit_harajat_amount'] = amount
        await update.message.reply_text("Yangi sanani kiriting (DD-MM-YYYY):", reply_markup=nav_keyboard())
        context.user_data['current_state'] = HARAJAT_EDIT_DATE
        return HARAJAT_EDIT_DATE
    except ValueError:
        await update.message.reply_text("Noto‘g‘ri miqdor. Qaytadan kiriting:", reply_markup=nav_keyboard())
        return HARAJAT_EDIT_AMOUNT

async def harajat_edit_date(update, context):
    date_text = update.message.text.strip()
    try:
        datetime.strptime(date_text, "%d-%m-%Y")
        context.user_data['edit_harajat_date'] = date_text
        await update.message.reply_text("Yangi sababni kiriting (ixtiyoriy, '-' deb yuboring):", reply_markup=nav_keyboard())
        context.user_data['current_state'] = HARAJAT_EDIT_REASON
        return HARAJAT_EDIT_REASON
    except ValueError:
        await update.message.reply_text("Sana formati noto‘g‘ri! To‘g‘ri format: DD-MM-YYYY.", reply_markup=nav_keyboard())
        return HARAJAT_EDIT_DATE

async def harajat_edit_reason(update, context):
    reason = update.message.text.strip()
    context.user_data['edit_harajat_reason'] = "" if reason == "-" else reason
    user = await sync_to_async(CustomUser.objects.get)(id=context.user_data.get('user_id'))
    # Set ishchi based on harajat_type, same logic for all users
    context.user_data['edit_harajat_ishchi'] = user if context.user_data.get('harajat_type') == 'ozimga' else None
    return await harajat_edit_confirm(update, context)

async def harajat_edit_ishchi(update, context):
    text = update.message.text.strip()
    if text == "-":
        context.user_data['edit_harajat_ishchi'] = None
        return await harajat_edit_confirm(update, context)
    users = context.user_data.get('users_list', [])
    if not users:
        users = await sync_to_async(list)(CustomUser.objects.all().order_by('first_name', 'last_name'))
        context.user_data['users_list'] = users
    if not users:
        await update.message.reply_text("Hech qanday ishchi topilmadi.", reply_markup=nav_keyboard())
        context.user_data['edit_harajat_ishchi'] = None
        return await harajat_edit_confirm(update, context)
    if text.isdigit():
        idx = int(text) - 1
        if 0 <= idx < len(users):
            context.user_data['edit_harajat_ishchi'] = users[idx]
            return await harajat_edit_confirm(update, context)
        else:
            await update.message.reply_text("Noto‘g‘ri raqam. Qaytadan tanlang yoki 'Ishxona' uchun '-' kiriting:", reply_markup=nav_keyboard())
            return HARAJAT_EDIT_ISHCHI
    await update.message.reply_text("Faqat ro‘yxatdan raqam kiriting yoki 'Ishxona' uchun '-' kiriting!", reply_markup=nav_keyboard())
    return HARAJAT_EDIT_ISHCHI

async def harajat_edit_confirm(update, context):
    if update.callback_query:
        query = update.callback_query
        await query.answer()
        if query.data == "confirm_yes":
            harajat = context.user_data.get('selected_harajat')
            user_data = context.user_data
            try:
                harajat.summa = user_data['edit_harajat_amount']
                harajat.sana = datetime.strptime(user_data['edit_harajat_date'], "%d-%m-%Y").date()
                harajat.sabab = user_data['edit_harajat_reason']
                harajat.ishchi = user_data.get('edit_harajat_ishchi')
                await sync_to_async(harajat.save)()
                await query.message.reply_text("Harajat yangilandi.", reply_markup=nav_keyboard())
                context.user_data['current_state'] = None
                return await show_harajat_list(update, context)
            except Exception as e:
                logger.error(f"Error updating expense: {e}")
                await query.message.reply_text("Xatolik yuz berdi. Qayta urinib ko‘ring.", reply_markup=nav_keyboard())
                return HARAJAT_EDIT_CONFIRM
        elif query.data == "confirm_no":
            context.user_data['current_state'] = HARAJAT_ACTION
            return await harajat_action_handler(update, context)
    else:
        user_data = context.user_data
        harajat = user_data.get('selected_harajat')
        kim = harajat.ishchi.first_name + " " + harajat.ishchi.last_name if harajat.ishchi else "Ishxona"
        text = (
            f"Tahrirlangan harajat:\n"
            f"Kim uchun: {kim}\n"
            f"Miqdor: {user_data['edit_harajat_amount']}\n"
            f"Sana: {user_data['edit_harajat_date']}\n"
            f"Sabab: {user_data['edit_harajat_reason'] or '-'}\n"
            f"Ma’lumotlarni tasdiqlaysizmi?"
        )
        keyboard = [
            [InlineKeyboardButton("✅ Ha", callback_data="confirm_yes"), InlineKeyboardButton("❌ Yo‘q", callback_data="confirm_no")]
        ]
        markup = nav_keyboard(extra=keyboard)
        await update.message.reply_text(text, reply_markup=markup)
        context.user_data['current_state'] = HARAJAT_EDIT_CONFIRM
        return HARAJAT_EDIT_CONFIRM

async def harajat_excel_menu(update, context):
    query = update.callback_query
    await query.answer()
    logger.info(f"harajat_excel_menu called with data: {query.data}")
    try:
        if query.data == "excel_full":
            context.user_data['excel_filter'] = ("full", None, None)
            logger.info("Exporting for full period")
            return await harajat_excel_export(update, context)
        elif query.data == "excel_interval":
            logger.info("Requesting date interval for export")
            await query.message.reply_text("Boshlanish sanasini kiriting (DD-MM-YYYY):", reply_markup=nav_keyboard())
            context.user_data['current_state'] = HARAJAT_EXCEL_DATE_START
            return HARAJAT_EXCEL_DATE_START
        else:
            logger.warning(f"Unhandled callback data in harajat_excel_menu: {query.data}")
            return HARAJAT_EXCEL_MENU
    except Exception as e:
        logger.error(f"Error in harajat_excel_menu: {e}", exc_info=True)
        await query.message.reply_text("Xatolik yuz berdi, qayta urinib ko‘ring.", reply_markup=nav_keyboard())
        return HARAJAT_EXCEL_MENU

async def harajat_excel_date_start(update, context):
    text = update.message.text.strip()
    try:
        start_date = datetime.strptime(text, "%d-%m-%Y").date()
        context.user_data['excel_start_date'] = start_date
        await update.message.reply_text("Tugash sanasini kiriting (DD-MM-YYYY):", reply_markup=nav_keyboard())
        context.user_data['current_state'] = HARAJAT_EXCEL_DATE_END
        return HARAJAT_EXCEL_DATE_END
    except ValueError:
        await update.message.reply_text("Sana formati noto‘g‘ri! To‘g‘ri format: DD-MM-YYYY.", reply_markup=nav_keyboard())
        return HARAJAT_EXCEL_DATE_START

async def harajat_excel_date_end(update, context):
    text = update.message.text.strip()
    try:
        end_date = datetime.strptime(text, "%d-%m-%Y").date()
        start_date = context.user_data.get('excel_start_date')
        if not start_date:
            await update.message.reply_text("Boshlanish sanasi topilmadi. Qaytadan urinib ko‘ring.", reply_markup=nav_keyboard())
            return HARAJAT_EXCEL_DATE_START
        if end_date < start_date:
            await update.message.reply_text("Tugash sanasi boshlanish sanasidan oldin bo‘lishi mumkin emas!", reply_markup=nav_keyboard())
            return HARAJAT_EXCEL_DATE_END
        context.user_data['excel_filter'] = ("interval", start_date, end_date)
        return await harajat_excel_export(update, context, start_date=start_date, end_date=end_date)
    except ValueError:
        await update.message.reply_text("Sana formati noto‘g‘ri! To‘g‘ri format: DD-MM-YYYY.", reply_markup=nav_keyboard())
        return HARAJAT_EXCEL_DATE_END
async def harajat_excel_export(update, context, start_date=None, end_date=None):
    from io import BytesIO
    from datetime import date
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font
    from openpyxl.cell.cell import MergedCell

    query = update.callback_query if hasattr(update, 'callback_query') else None
    if not await check_login(context, update):
        return ConversationHandler.END

    if not start_date or not end_date:
        excel_filter = context.user_data.get('excel_filter')
        if excel_filter and excel_filter[0] == "interval":
            start_date, end_date = excel_filter[1], excel_filter[2]
        else:
            try:
                earliest = await sync_to_async(Harajatlar.objects.earliest)('sana')
                latest = await sync_to_async(Harajatlar.objects.latest)('sana')
                start_date, end_date = earliest.sana, latest.sana
            except Harajatlar.DoesNotExist:
                start_date = date.today()
                end_date = date.today()

    filter_kwargs = {'sana__gte': start_date, 'sana__lte': end_date}
    harajatlar = await sync_to_async(list)(
        Harajatlar.objects.filter(**filter_kwargs).order_by('sana')
    )

    # Main sheet rows (all expenses)
    rows = []
    for harajat in harajatlar:
        if harajat.ishchi_id:
            ishchi = await sync_to_async(lambda: harajat.ishchi)()
            kim = f"{ishchi.first_name} {ishchi.last_name}"
        else:
            kim = "Ishxona"
        rows.append({
            "Sana": harajat.sana.strftime("%d-%m-%Y"),
            "Kim uchun": kim,
            "Summa": float(harajat.summa),
            "Sabab": harajat.sabab or "-"
        })
    jami_summa = sum(row["Summa"] for row in rows)
    if rows:
        rows.append({
            "Sana": "",
            "Kim uchun": "",
            "Summa": jami_summa,
            "Sabab": "Jami"
        })
    df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=["Sana", "Kim uchun", "Summa", "Sabab"])

    # User-specific sheets
    all_users = await sync_to_async(list)(CustomUser.objects.all())
    user_dfs = {}
    for user in all_users:
        user_rows = [
            {
                "Sana": harajat.sana.strftime("%d-%m-%Y"),
                "Summa": float(harajat.summa),
                "Sabab": harajat.sabab or "-"
            }
            for harajat in harajatlar if harajat.ishchi_id == user.id
        ]
        user_jami = sum(row["Summa"] for row in user_rows)
        if user_rows:
            user_rows.append({
                "Sana": "",
                "Summa": user_jami,
                "Sabab": "Jami"
            })
        user_dfs[user] = pd.DataFrame(user_rows) if user_rows else pd.DataFrame(columns=["Sana", "Summa", "Sabab"])

    # "Ishxona" sheet (expenses with no user)
    ishxona_rows = [
        {
            "Sana": harajat.sana.strftime("%d-%m-%Y"),
            "Summa": float(harajat.summa),
            "Sabab": harajat.sabab or "-"
        }
        for harajat in harajatlar if not harajat.ishchi_id
    ]
    ishxona_jami = sum(row["Summa"] for row in ishxona_rows)
    if ishxona_rows:
        ishxona_rows.append({
            "Sana": "",
            "Summa": ishxona_jami,
            "Sabab": "Jami"
        })
    ishxona_df = pd.DataFrame(ishxona_rows) if ishxona_rows else pd.DataFrame(columns=["Sana", "Summa", "Sabab"])

    # === Excel export ===
    with BytesIO() as output:
        workbook = Workbook()
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        # Main sheet
        main_sheet = workbook.create_sheet('Harajatlar Hisoboti')
        main_sheet.merge_cells('A1:D1')
        title_cell = main_sheet['A1']
        title_cell.value = f"Harajatlar Hisoboti: {start_date} dan {end_date} gacha"
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.font = Font(bold=True)
        for col_idx, col_name in enumerate(df.columns, start=1):
            main_sheet[f"{get_column_letter(col_idx)}2"] = col_name
            main_sheet[f"{get_column_letter(col_idx)}2"].font = Font(bold=True)
        for row_idx, row in enumerate(df.values, start=3):
            for col_idx, value in enumerate(row, start=1):
                cell = main_sheet[f"{get_column_letter(col_idx)}{row_idx}"]
                cell.value = value
                if row_idx == len(df) + 2 and rows:  # Bold total row
                    cell.font = Font(bold=True)
        for col_idx in range(1, len(df.columns) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(str(df.columns[col_idx-1]))
            for row_idx in range(2, main_sheet.max_row + 1):
                cell = main_sheet[f"{column_letter}{row_idx}"]
                if not isinstance(cell, MergedCell):
                    max_length = max(max_length, len(str(cell.value or "")))
            main_sheet.column_dimensions[column_letter].width = max_length + 2

        # Each user's sheet
        for user, user_df in user_dfs.items():
            user_name = f"{user.first_name or ''} {user.last_name or ''}".strip() or user.username or f"user{user.id}"
            sheet_name = user_name[:31].replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_')
            user_sheet = workbook.create_sheet(sheet_name)
            user_sheet.merge_cells('A1:C1')
            user_title = user_sheet['A1']
            user_title.value = f"{user_name} harajatlari: {start_date} dan {end_date} gacha"
            user_title.alignment = Alignment(horizontal='center', vertical='center')
            user_title.font = Font(bold=True)
            for col_idx, col_name in enumerate(user_df.columns, start=1):
                user_sheet[f"{get_column_letter(col_idx)}2"] = col_name
                user_sheet[f"{get_column_letter(col_idx)}2"].font = Font(bold=True)
            for row_idx, row in enumerate(user_df.values, start=3):
                for col_idx, value in enumerate(row, start=1):
                    cell = user_sheet[f"{get_column_letter(col_idx)}{row_idx}"]
                    cell.value = value
                    if row_idx == len(user_df) + 2 and not user_df.empty:
                        cell.font = Font(bold=True)
            for col_idx in range(1, len(user_df.columns) + 1):
                column_letter = get_column_letter(col_idx)
                max_length = len(str(user_df.columns[col_idx-1]))
                for row_idx in range(2, user_sheet.max_row + 1):
                    cell = user_sheet[f"{column_letter}{row_idx}"]
                    if not isinstance(cell, MergedCell):
                        max_length = max(max_length, len(str(cell.value or "")))
                user_sheet.column_dimensions[column_letter].width = max_length + 2

        # Ishxona sheet
        ishxona_sheet = workbook.create_sheet('Ishxona')
        ishxona_sheet.merge_cells('A1:C1')
        ishxona_title = ishxona_sheet['A1']
        ishxona_title.value = f"Ishxona harajatlari: {start_date} dan {end_date} gacha"
        ishxona_title.alignment = Alignment(horizontal='center', vertical='center')
        ishxona_title.font = Font(bold=True)
        for col_idx, col_name in enumerate(ishxona_df.columns, start=1):
            ishxona_sheet[f"{get_column_letter(col_idx)}2"] = col_name
            ishxona_sheet[f"{get_column_letter(col_idx)}2"].font = Font(bold=True)
        for row_idx, row in enumerate(ishxona_df.values, start=3):
            for col_idx, value in enumerate(row, start=1):
                cell = ishxona_sheet[f"{get_column_letter(col_idx)}{row_idx}"]
                cell.value = value
                if row_idx == len(ishxona_df) + 2 and not ishxona_df.empty:
                    cell.font = Font(bold=True)
        for col_idx in range(1, len(ishxona_df.columns) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(str(ishxona_df.columns[col_idx-1]))
            for row_idx in range(2, ishxona_sheet.max_row + 1):
                cell = ishxona_sheet[f"{column_letter}{row_idx}"]
                if not isinstance(cell, MergedCell):
                    max_length = max(max_length, len(str(cell.value or "")))
            ishxona_sheet.column_dimensions[column_letter].width = max_length + 2

        workbook.save(output)
        output.seek(0)
        await (query.message if query else update.message).reply_document(
            document=output, filename=f"Harajatlar_{start_date}_to_{end_date}.xlsx"
        )
    context.user_data['current_state'] = None
    await show_menu(update, context)
    return ConversationHandler.END

# Qolgan Harajatlar handler’lari:
# - harajat_date, harajat_reason, harajat_confirmation
# - show_harajat_list, harajat_action, harajat_delete_confirm
# - tahrirlash funksiyalari (HARAJAT_EDIT_*)
# - Excel eksport (HARAJAT_EXCEL_*)




# --- YANGI TAMINOTCHI DIALOGI ---
async def inline_callback(update, context):
    query = update.callback_query
    await query.answer()
    data = query.data
    logger.info(f"inline_callback called with data: {data}, current_state: {context.user_data.get('current_state')}")
    try:
        if data == "btn_harajatlar":
            logger.info("Harajatlar button clicked")
            keyboard = [
                [InlineKeyboardButton("O‘zimga", callback_data="harajat_ozimga")],
                [InlineKeyboardButton("Ishxonaga", callback_data="harajat_ishxonaga")],
                [InlineKeyboardButton("Ishchilarga", callback_data="harajat_ishchilarga")],
                [InlineKeyboardButton("Ro‘yxat", callback_data="harajat_list")]
            ]
            markup = nav_keyboard(extra=keyboard)
            await query.message.reply_text("Harajatlar:", reply_markup=markup)
            context.user_data['current_state'] = HARAJAT_MENU
            return HARAJAT_MENU
        elif query.data == "btn_taminotchi":
            context.user_data["add_supplier_context"] = True
            await query.message.reply_text("Taminotchi ismini kiriting:", reply_markup=nav_keyboard())
            context.user_data['current_state'] = SUPPLIER_NAME
            return SUPPLIER_NAME
        elif query.data == "btn_hisobot":
            keyboard = [
                [InlineKeyboardButton("Butun davr", callback_data="report_full")],
                [InlineKeyboardButton("Ma’lum davr", callback_data="report_interval")],
            ]
            markup = nav_keyboard(extra=keyboard)
            await query.message.reply_text("Hisobot qaysi davr uchun kerak?", reply_markup=markup)
            context.user_data['current_state'] = REPORT_MENU
            return REPORT_MENU
        elif data == "btn_ishchilar":
            logger.info("Ishchilar button clicked")
            user = await sync_to_async(CustomUser.objects.get)(id=context.user_data.get('user_id'))
            if not user.can_add_new_users:
                logger.info(f"User {user.id} lacks can_add_new_users permission")
                await query.message.reply_text("Sizda ishchilar bilan ishlash huquqi yo‘q.", reply_markup=nav_keyboard())
                return ConversationHandler.END
            context.user_data['ishchi_page'] = 1
            return await show_ishchilar(update, context)
        else:
            logger.warning(f"Unhandled callback data: {data}")
            await query.message.reply_text("Noto‘g‘ri buyruq, qaytadan urinib ko‘ring.", reply_markup=nav_keyboard())
            return ConversationHandler.END
    except Exception as e:
        logger.error(f"Error in inline_callback: {e}", exc_info=True)
        await query.message.reply_text("Xatolik yuz berdi, qayta urinib ko‘ring.", reply_markup=nav_keyboard())
        return ConversationHandler.END

async def supplier_name(update, context):
    context.user_data["supplier_name"] = update.message.text.strip()
    await update.message.reply_text("Taminotchi telefon raqamini kiriting (masalan: +998991234567):", reply_markup=nav_keyboard())
    context.user_data['current_state'] = SUPPLIER_PHONE
    return SUPPLIER_PHONE

async def supplier_phone(update, context):
    supplier_name = context.user_data.get("supplier_name")
    supplier_phone = update.message.text.strip()
    try:
        supplier = await sync_to_async(Taminotchi.objects.create)(
            taminotchi_ismi=supplier_name,
            taminotchi_telefon_raqami=supplier_phone
        )
        await update.message.reply_text(f"Taminotchi {supplier_name} qo‘shildi!", reply_markup=nav_keyboard())
        if context.user_data.get("add_supplier_context"):
            context.user_data.pop("add_supplier_context")
            context.user_data.pop('current_state', None)
            context.user_data.pop('supplier_name', None)
            await show_menu(update, context)
            return ConversationHandler.END
        context.user_data["selected_supplier"] = supplier
        context.user_data.pop('current_state', None)
        context.user_data.pop('supplier_name', None)
        return await show_supplier_info(update, context, supplier)
    except Exception as e:
        logger.error(f"Error in supplier_phone: {e}")
        await update.message.reply_text("Xatolik yuz berdi. Qayta urinib ko‘ring.", reply_markup=nav_keyboard())
        return SUPPLIER_PHONE

# --- TAMINOTCHILAR RO'YXATI, TANLASH, QIDIRUV, INFO ---
async def show_suppliers(update, context):
    if not await check_login(context, update):
        return ConversationHandler.END
    page = context.user_data.get('supplier_page', 1)
    suppliers = await sync_to_async(list)(Taminotchi.objects.all())
    if not suppliers:
        text = "Hech qanday taminotchi topilmadi."
        if update.callback_query:
            await update.callback_query.message.reply_text(text, reply_markup=nav_keyboard())
        else:
            await update.message.reply_text(text, reply_markup=nav_keyboard())
        context.user_data['current_state'] = CHOOSE_SUPPLIER
        return CHOOSE_SUPPLIER
    total = len(suppliers)
    start = (page - 1) * PAGE_SIZE
    end = start + PAGE_SIZE
    page_suppliers = suppliers[start:end]
    context.user_data['suppliers_list'] = suppliers
    text = f"Taminotchilar ro'yxati ({start+1}-{min(end, total)}/Jami: {total}):\n"
    for idx, s in enumerate(page_suppliers, start=1):
        umumiy_qarz = await sync_to_async(s.umumiy_qarz)()
        text += f"{start+idx}. {s.taminotchi_ismi} (Tel: {s.taminotchi_telefon_raqami or '-'}, qarz: {umumiy_qarz})\n"
    row = []
    if start > 0:
        row.append(InlineKeyboardButton("⬅️ Oldingi", callback_data="prev_suppliers"))
    if end < total:
        row.append(InlineKeyboardButton("➡️ Keyingi", callback_data="next_suppliers"))
    keyboard = []
    if row: keyboard.append(row)
    keyboard.append([
        InlineKeyboardButton("Qidirish", callback_data="find_supplier"),
        InlineKeyboardButton("Qo‘shish", callback_data="add_supplier"),
    ])
    markup = nav_keyboard(extra=keyboard)
    if update.callback_query:
        await update.callback_query.edit_message_text(text, reply_markup=markup)
    else:
        await update.message.reply_text(text, reply_markup=markup)
    context.user_data['current_state'] = CHOOSE_SUPPLIER
    return CHOOSE_SUPPLIER

async def taminotchilar_callback(update, context):
    query = update.callback_query
    await query.answer()
    if not await check_login(context, update):
        return ConversationHandler.END
    data = query.data
    logger.info(f"taminotchilar_callback: data={data}, current_state={context.user_data.get('current_state')}")
    if data == "btn_taminotchilar":
        context.user_data['supplier_page'] = 1
        return await show_suppliers(update, context)
    elif data == "prev_suppliers":
        context.user_data['supplier_page'] = max(context.user_data.get('supplier_page', 1) - 1, 1)
        return await show_suppliers(update, context)
    elif data == "next_suppliers":
        context.user_data['supplier_page'] = context.user_data.get('supplier_page', 1) + 1
        return await show_suppliers(update, context)
    elif data == "add_supplier":
        context.user_data["add_supplier_context"] = False
        await query.message.reply_text("Yangi taminotchi ismini kiriting:", reply_markup=nav_keyboard())
        context.user_data['current_state'] = SUPPLIER_NAME
        return SUPPLIER_NAME
    elif data == "find_supplier":
        await query.message.reply_text("Qidiruv uchun ismni yozing:", reply_markup=nav_keyboard())
        context.user_data['current_state'] = FIND_SUPPLIER
        return FIND_SUPPLIER

async def choose_supplier(update, context):
    text = update.message.text.strip()
    suppliers = context.user_data.get('suppliers_list', [])
    if not suppliers:
        suppliers = await sync_to_async(list)(Taminotchi.objects.all())
        context.user_data['suppliers_list'] = suppliers
    if not suppliers:
        await update.message.reply_text("Hech qanday taminotchi topilmadi.", reply_markup=nav_keyboard())
        context.user_data['current_state'] = CHOOSE_SUPPLIER
        return CHOOSE_SUPPLIER
    if text.isdigit():
        idx = int(text) - 1
        if 0 <= idx < len(suppliers):
            supplier = suppliers[idx]
            context.user_data["selected_supplier"] = supplier
            context.user_data['current_state'] = None
            return await show_supplier_info(update, context, supplier)
        else:
            await update.message.reply_text("Xato raqam. Qaytadan tanlang:", reply_markup=nav_keyboard())
            return CHOOSE_SUPPLIER
    await update.message.reply_text("Faqat ro‘yxatdan raqam kiriting!", reply_markup=nav_keyboard())
    return CHOOSE_SUPPLIER

async def find_supplier(update, context):
    search_text = update.message.text.strip()
    suppliers = await sync_to_async(list)(
        Taminotchi.objects.filter(taminotchi_ismi__icontains=search_text)
    )
    if not suppliers:
        await update.message.reply_text("Hech qanday taminotchi topilmadi.", reply_markup=nav_keyboard())
        context.user_data['current_state'] = CHOOSE_SUPPLIER
        return CHOOSE_SUPPLIER
    context.user_data['suppliers_list'] = suppliers
    text = "Qidiruv natijalari:\n"
    for i, s in enumerate(suppliers, start=1):
        text += f"{i}. {s.taminotchi_ismi} ({s.taminotchi_telefon_raqami or '-'})\n"
    text += "\nIltimos, kerakli taminotchining raqamini kiriting:"
    await update.message.reply_text(text, reply_markup=nav_keyboard())
    context.user_data['current_state'] = CHOOSE_SUPPLIER
    return CHOOSE_SUPPLIER

# --- TAMINOTCHI INFO + 3TA TUGMA ---
async def show_supplier_info(update, context, supplier):
    context.user_data["selected_supplier"] = supplier
    context.user_data["selected_supplier_id"] = supplier.id
    umumiy_qarz = await sync_to_async(supplier.umumiy_qarz)()
    text = (
        f"Tanlangan taminotchi:\n"
        f"Ism: {supplier.taminotchi_ismi}\n"
        f"Telefon: {supplier.taminotchi_telefon_raqami or '-'}\n"
        f"Qarzi: {umumiy_qarz}\n"
    )
    keyboard = [
        [
            InlineKeyboardButton("Qarz bo'lish", callback_data="btn_qarz_bolish"),
            InlineKeyboardButton("Qarzni to‘lash", callback_data="btn_qarzni_tolash")
        ],
        [InlineKeyboardButton("📥 Excel", callback_data=f"excel_{supplier.id}")],
    ]
    markup = nav_keyboard(extra=keyboard)
    if hasattr(update, "callback_query") and update.callback_query:
        await update.callback_query.message.reply_text(text, reply_markup=markup)
    else:
        await update.message.reply_text(text, reply_markup=markup)
    context.user_data['current_state'] = SUPPLIER_ACTION
    return SUPPLIER_ACTION

async def supplier_action_handler(update, context):
    query = update.callback_query
    await query.answer()
    if not await check_login(context, update):
        return ConversationHandler.END
    logger.info(f"supplier_action_handler: data={query.data}, current_state={context.user_data.get('current_state')}")
    if query.data == "btn_qarz_bolish":
        keyboard = [
            [InlineKeyboardButton("So‘m", callback_data="currency_sum")],
            [InlineKeyboardButton("Dollar", callback_data="currency_usd")],
        ]
        markup = InlineKeyboardMarkup(keyboard)
        await query.message.reply_text("To‘lov valyutasini tanlang:", reply_markup=markup)
        context.user_data['current_state'] = PAYMENT_CURRENCY
        return PAYMENT_CURRENCY
    elif query.data == "btn_qarzni_tolash":
        keyboard = [
            [InlineKeyboardButton("So‘m", callback_data="currency_sum")],
            [InlineKeyboardButton("Dollar", callback_data="currency_usd")],
        ]
        markup = InlineKeyboardMarkup(keyboard)
        await query.message.reply_text("To‘lov valyutasini tanlang:", reply_markup=markup)
        context.user_data['current_state'] = ASK_PAYMENT_CURRENCY
        return ASK_PAYMENT_CURRENCY


    elif query.data == "btn_back":
        context.user_data['current_state'] = None
        return await show_suppliers(update, context)
    elif query.data == "btn_menu":
        context.user_data['current_state'] = None
        await show_menu(update, context)
        return ConversationHandler.END

# --- QARZ OQIMI ---
async def ask_currency(update, context):
    query = update.callback_query
    await query.answer()
    currency = query.data

    if currency == "currency_sum":
        context.user_data['currency'] = "UZS"
    elif currency == "currency_usd":
        context.user_data['currency'] = "USD"
    else:
        await query.message.reply_text("Noto‘g‘ri tanlov. Qayta urinib ko‘ring.")
        return PAYMENT_CURRENCY

    await query.message.reply_text("Olingan summani kiriting:", reply_markup=nav_keyboard())
    context.user_data['current_state'] = ASK_AMOUNT
    return ASK_AMOUNT
async def ask_amount(update, context):
    try:
        amount = float(update.message.text.replace(",", "."))
        if amount <= 0:
            await update.message.reply_text("Miqdor musbat bo‘lishi kerak. Qaytadan kiriting:", reply_markup=nav_keyboard())
            return ASK_AMOUNT
        context.user_data["qarz_miqdori"] = amount
        await update.message.reply_text("Qarz sanasini kiriting (DD-MM-YYYY):", reply_markup=nav_keyboard())
        context.user_data['current_state'] = ASK_DATE
        return ASK_DATE
    except ValueError:
        await update.message.reply_text("Noto‘g‘ri miqdor. Qaytadan kiriting:", reply_markup=nav_keyboard())
        return ASK_AMOUNT

async def ask_date(update, context):
    date_text = update.message.text.strip()
    try:
        datetime.strptime(date_text, "%d-%m-%Y")
        context.user_data["qarz_sanasi"] = date_text
        await update.message.reply_text("Sababni kiriting (ixtiyoriy, '-' deb yuboring):", reply_markup=nav_keyboard())
        context.user_data['current_state'] = ASK_REASON
        return ASK_REASON
    except ValueError:
        await update.message.reply_text("Sana formati noto‘g‘ri! To‘g‘ri format: DD-MM-YYYY.", reply_markup=nav_keyboard())
        return ASK_DATE

async def ask_reason(update, context):
    reason = update.message.text.strip()
    if reason == "-":
        reason = ""
    supplier = context.user_data.get("selected_supplier")
    if not supplier:
        await update.message.reply_text("Taminotchi tanlanmadi. Qaytadan boshlang.", reply_markup=nav_keyboard())
        context.user_data['current_state'] = None
        await show_menu(update, context)
        return ConversationHandler.END
    miqdor = context.user_data.get("qarz_miqdori")
    sana = context.user_data.get("qarz_sanasi")
    text = (
        "Kiritilgan ma’lumotlar:\n"
        f"Taminotchi: {supplier.taminotchi_ismi}\n"
        f"Qarz miqdori: {miqdor}\n"
        f"Sana: {sana}\n"
        f"Sabab: {reason or '-'}\n"
        "Ma’lumotlar to‘g‘rimi?"
    )
    confirm_buttons = [
        InlineKeyboardButton("✅ Ha", callback_data="btn_confirm"),
        InlineKeyboardButton("❌ Yo‘q, boshidan", callback_data="btn_reject")
    ]
    markup = nav_keyboard(extra=confirm_buttons)
    await update.message.reply_text(text, reply_markup=markup)
    context.user_data['qarz_sababi'] = reason
    context.user_data['current_state'] = CONFIRMATION
    return CONFIRMATION

async def confirmation_handler(update, context):
    query = update.callback_query
    await query.answer()
    logger.info(f"confirmation_handler: data={query.data}, current_state={context.user_data.get('current_state')}")
    if query.data == "btn_confirm":
        supplier = context.user_data.get('selected_supplier')
        if not supplier:
            await query.message.reply_text("Taminotchi tanlanmadi. Qaytadan boshlang.", reply_markup=nav_keyboard())
            context.user_data['current_state'] = None
            await show_menu(update, context)
            return ConversationHandler.END
        miqdor = context.user_data.get('qarz_miqdori')
        sana = context.user_data.get('qarz_sanasi')
        sana = datetime.strptime(sana, "%d-%m-%Y").date()
        reason = context.user_data.get('qarz_sababi')
        currency = context.user_data.get('currency')

        try:
            exists = await sync_to_async(Pul_olish.objects.filter(taminotchi=supplier).exists)()
            umumiy_qarz = await sync_to_async(supplier.umumiy_qarz)()
            miqdor = Decimal(str(miqdor))  # Har doim Decimal!
            if umumiy_qarz < 0:
                umumiy_qarz = umumiy_qarz * (-1)
                if umumiy_qarz - miqdor >= 0:
                
                    qoldiq = miqdor
                    status = "tolangan"
                else:
                    qoldiq = umumiy_qarz
                    status = "tolanmagan"
                b = await sync_to_async(Pul_olish.objects.create)(
                    taminotchi=supplier,
                    sabab=reason,
                    sana=sana,
                    umumiy_miqdor=miqdor,
                    tolangan=qoldiq,
                    status = status,
                    currency=currency
                )

                
            else:
                b = await sync_to_async(Pul_olish.objects.create)(
                taminotchi=supplier, sabab=reason, sana=sana, umumiy_miqdor=miqdor,currency=currency
                )
            
            if not exists:
                await sync_to_async(Pul_berish.objects.create)(
                    taminotchi=supplier,
                    summa=0,
                    sana=sana,
                    notification_sent=True,
                    berildi=True,
                    currency=currency,
                    pul_olingan=b
                )
            await query.message.reply_text("Qarz ma’lumoti saqlandi!", reply_markup=nav_keyboard())
            context.user_data['current_state'] = None
            await show_menu(update, context)
            return ConversationHandler.END
        except Exception as e:
            logger.error(f"Error in confirmation_handler: {e}")
            await query.message.reply_text("Xatolik yuz berdi. Qayta urinib ko‘ring.", reply_markup=nav_keyboard())
            return CONFIRMATION
    elif query.data == "btn_reject":
        context.user_data['current_state'] = None
        await show_menu(update, context)
        return ConversationHandler.END
    elif query.data == "btn_menu":
        context.user_data['current_state'] = None
        await show_menu(update, context)
        return ConversationHandler.END
    elif query.data == "btn_back":
        await query.message.reply_text("Sababni kiriting (ixtiyoriy, '-' deb yuboring):", reply_markup=nav_keyboard())
        context.user_data['current_state'] = ASK_REASON
        return ASK_REASON

# --- NAVIGATION HANDLER ---
async def navigation_callback(update, context):
    query = update.callback_query
    await query.answer()
    state = context.user_data.get('current_state', None)
    logger.info(f"navigation_callback: data={query.data}, current_state={state}")
    if query.data == "btn_menu":
        context.user_data['current_state'] = None
        await show_menu(update, context)
        return ConversationHandler.END
    elif query.data == "btn_back":
        if state == ASK_LOGIN:
            context.user_data['current_state'] = None
            await show_menu(update, context)
            return ConversationHandler.END
        elif state == CHOOSE_ISHCHI:  # NEW
            context.user_data['current_state'] = None
            await show_menu(update, context)
            return ConversationHandler.END
        elif state == ISHCHI_NAME:  # NEW
            context.user_data['current_state'] = None
            await show_menu(update, context)
            return ConversationHandler.END
        elif state == ISHCHI_SURNAME:  # NEW
            await query.message.reply_text("Yangi ishchi ismini kiriting:", reply_markup=nav_keyboard())
            context.user_data['current_state'] = ISHCHI_NAME
            return ISHCHI_NAME
        elif state == ISHCHI_PHONE:  # NEW
            await query.message.reply_text("Familiyasini kiriting:", reply_markup=nav_keyboard())
            context.user_data['current_state'] = ISHCHI_SURNAME
            return ISHCHI_SURNAME
        elif state == ISHCHI_PERMISSIONS:  # NEW
            await query.message.reply_text("Telefon raqamini kiriting (masalan, +998991234567):", reply_markup=nav_keyboard())
            context.user_data['current_state'] = ISHCHI_PHONE
            return ISHCHI_PHONE
        elif state == ISHCHI_CONFIRM:  # NEW
            perm_step = context.user_data.get('perm_step', 'can_add_expanse')
            if perm_step == 'can_add_expanse':
                await query.message.reply_text("Telefon raqamini kiriting (masalan, +998991234567):", reply_markup=nav_keyboard())
                context.user_data['current_state'] = ISHCHI_PHONE
                return ISHCHI_PHONE
            elif perm_step == 'can_add_expanse_to_others':
                keyboard = [[InlineKeyboardButton("Ha", callback_data="perm_yes"), InlineKeyboardButton("Yo‘q", callback_data="perm_no")]]
                markup = nav_keyboard(extra=keyboard)
                await query.message.reply_text("Harajat qo‘shish ruxsatini berasizmi?", reply_markup=markup)
                context.user_data['current_state'] = ISHCHI_PERMISSIONS
                context.user_data['perm_step'] = 'can_add_expanse'
                return ISHCHI_PERMISSIONS
            # Boshqa ruxsatlar uchun shunga o‘xshash
        elif state == FIND_ISHCHI:  # NEW
            context.user_data['current_state'] = None
            return await show_ishchilar(update, context)
        elif state == ISHCHI_ACTION:  # NEW
            context.user_data['current_state'] = None
            return await show_ishchilar(update, context)
        elif state == ISHCHI_DELETE_CONFIRM:  # NEW
            return await show_ishchi_info(update, context)
        elif state == ISHCHI_EDIT_NAME:  # NEW
            return await show_ishchi_info(update, context)
        elif state == ISHCHI_EDIT_SURNAME:  # NEW
            await query.message.reply_text("Yangi ismini kiriting:", reply_markup=nav_keyboard())
            context.user_data['current_state'] = ISHCHI_EDIT_NAME
            return ISHCHI_EDIT_NAME
        # Harajatlar uchun
        elif state == HARAJAT_MENU:  # NEW
            context.user_data['current_state'] = None
            await show_menu(update, context)
            return ConversationHandler.END
        elif state == HARAJAT_AMOUNT:  # NEW
            user = await sync_to_async(CustomUser.objects.get)(id=context.user_data.get('user_id'))
            keyboard = []
            if user.can_add_expanse:
                keyboard.append([
                    InlineKeyboardButton("O‘zimga", callback_data="harajat_ozimga"),
                    InlineKeyboardButton("Ishxonaga", callback_data="harajat_ishxonaga"),
                ])
            if user.can_add_expanse_to_others:
                keyboard.append([InlineKeyboardButton("Ishchilarga", callback_data="harajat_ishchilarga")])
            keyboard.append([InlineKeyboardButton("Harajatlar ro‘yxati", callback_data="harajat_list")])
            markup = nav_keyboard(extra=keyboard)
            await query.message.reply_text("Harajat turini tanlang:", reply_markup=markup)
            context.user_data['current_state'] = HARAJAT_MENU
            return HARAJAT_MENU
        elif state == HARAJAT_DATE:  # NEW
            await query.message.reply_text("Harajat miqdorini kiriting:", reply_markup=nav_keyboard())
            context.user_data['current_state'] = HARAJAT_AMOUNT
            return HARAJAT_AMOUNT
        elif state == HARAJAT_REASON:  # NEW
            await query.message.reply_text("Harajat sanasini kiriting (DD-MM-YYYY):", reply_markup=nav_keyboard())
            context.user_data['current_state'] = HARAJAT_DATE
            return HARAJAT_DATE
        elif state == CHOOSE_SUPPLIER:
            context.user_data['current_state'] = None
            await show_menu(update, context)
            return ConversationHandler.END
        elif state == SUPPLIER_NAME:
            context.user_data['current_state'] = None
            await show_menu(update, context)
            return ConversationHandler.END
        elif state == SUPPLIER_PHONE:
            await query.message.reply_text("Taminotchi ismini kiriting:", reply_markup=nav_keyboard())
            context.user_data['current_state'] = SUPPLIER_NAME
            return SUPPLIER_NAME
        elif state == FIND_SUPPLIER:
            context.user_data['current_state'] = None
            return await show_suppliers(update, context)
        elif state == SUPPLIER_ACTION:
            context.user_data['current_state'] = None
            return await show_suppliers(update, context)
        elif state == ASK_AMOUNT:
            supplier = context.user_data.get("selected_supplier")
            if supplier:
                return await show_supplier_info(update, context, supplier)
            context.user_data['current_state'] = None
            await show_menu(update, context)
            return ConversationHandler.END
        elif state == ASK_DATE:
            await query.message.reply_text("Qarz miqdorini kiriting:", reply_markup=nav_keyboard())
            context.user_data['current_state'] = ASK_AMOUNT
            return ASK_AMOUNT
        elif state == ASK_REASON:
            await query.message.reply_text("Qarz sanasini kiriting (DD-MM-YYYY):", reply_markup=nav_keyboard())
            context.user_data['current_state'] = ASK_DATE
            return ASK_DATE
        elif state == CONFIRMATION:
            await query.message.reply_text("Sababni kiriting (ixtiyoriy, '-' deb yuboring):", reply_markup=nav_keyboard())
            context.user_data['current_state'] = ASK_REASON
            return ASK_REASON
        elif state == ASK_PAYMENT_AMOUNT:
            supplier = context.user_data.get("selected_supplier")
            if supplier:
                return await show_supplier_info(update, context, supplier)
            context.user_data['current_state'] = None
            await show_menu(update, context)
            return ConversationHandler.END
        elif state == ASK_PAYMENT_DATE:
            await query.message.reply_text("To‘lanadigan summani kiriting:", reply_markup=nav_keyboard())
            context.user_data['current_state'] = ASK_PAYMENT_AMOUNT
            return ASK_PAYMENT_AMOUNT
        elif state == PAYMENT_CONFIRM:
            await query.message.reply_text("To‘lov sanasini kiriting (DD-MM-YYYY):", reply_markup=nav_keyboard())
            context.user_data['current_state'] = ASK_PAYMENT_DATE
            return ASK_PAYMENT_DATE
        elif state == EXCEL_MENU:
            supplier = context.user_data.get("selected_supplier")
            if supplier:
                return await show_supplier_info(update, context, supplier)
            context.user_data['current_state'] = None
            await show_menu(update, context)
            return ConversationHandler.END
        elif state == EXCEL_DATE_START:
            await query.message.reply_text("Ma’lumot qaysi davr uchun kerak?", reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("Butun davr", callback_data="excel_full")],
                [InlineKeyboardButton("Vaqt oralig‘i bo‘yicha", callback_data="excel_interval")],
                [InlineKeyboardButton("Ortga", callback_data="btn_back")],
                [InlineKeyboardButton("Menu", callback_data="btn_menu")],
            ]))
            context.user_data['current_state'] = EXCEL_MENU
            return EXCEL_MENU
        elif state == EXCEL_DATE_END:
            await query.message.reply_text("Boshlanish sanasini kiriting (DD-MM-YYYY):", reply_markup=nav_keyboard())
            context.user_data['current_state'] = EXCEL_DATE_START
            return EXCEL_DATE_START
        elif state == REPORT_MENU:
            context.user_data['current_state'] = None
            await show_menu(update, context)
            return ConversationHandler.END
        elif state == REPORT_DATE_START:
            context.user_data['current_state'] = None
            await show_menu(update, context)
            return ConversationHandler.END
        elif state == REPORT_DATE_END:
            await query.message.reply_text("Boshlanish sanasini kiriting (DD-MM-YYYY):", reply_markup=nav_keyboard())
            context.user_data['current_state'] = REPORT_DATE_START
            return REPORT_DATE_START
        else:
            context.user_data['current_state'] = None
            await show_menu(update, context)
            return ConversationHandler.END

# --- QARZNI TO'LASH ---
async def ask_payment_currency(update, context):
    query = update.callback_query
    await query.answer()
    currency = query.data

    if currency == "currency_sum":
        context.user_data['currency'] = "UZS"
    elif currency == "currency_usd":
        context.user_data['currency'] = "USD"
    else:
        await query.message.reply_text("Noto‘g‘ri tanlov. Qayta urinib ko‘ring.")
        return ASK_PAYMENT_CURRENCY

    await query.message.reply_text("To‘lanadigan summani kiriting:", reply_markup=nav_keyboard())
    context.user_data['current_state'] = ASK_PAYMENT_AMOUNT
    return ASK_PAYMENT_AMOUNT

async def ask_payment_amount(update, context):
    try:
        amount = float(update.message.text.replace(",", "."))
        if amount <= 0:
            await update.message.reply_text("Miqdor musbat bo‘lishi kerak. Qaytadan kiriting:", reply_markup=nav_keyboard())
            return ASK_PAYMENT_AMOUNT
        context.user_data["tolov_miqdori"] = amount
        await update.message.reply_text("To‘lov sanasini kiriting (DD-MM-YYYY):", reply_markup=nav_keyboard())
        context.user_data['current_state'] = ASK_PAYMENT_DATE
        return ASK_PAYMENT_DATE
    except ValueError:
        await update.message.reply_text("Noto‘g‘ri miqdor. Qaytadan kiriting:", reply_markup=nav_keyboard())
        return ASK_PAYMENT_AMOUNT

async def ask_payment_date(update, context):
    date_text = update.message.text.strip()
    try:
        datetime.strptime(date_text, "%d-%m-%Y")
        context.user_data["tolov_sanasi"] = date_text
        supplier = context.user_data.get("selected_supplier")
        if not supplier:
            await update.message.reply_text("Taminotchi tanlanmadi. Qaytadan boshlang.", reply_markup=nav_keyboard())
            context.user_data['current_state'] = None
            await show_menu(update, context)
            return ConversationHandler.END
        miqdor = context.user_data.get("tolov_miqdori")
        sana = context.user_data.get("tolov_sanasi")
        text = (
            f"To‘lov ma’lumotlari:\n"
            f"Taminotchi: {supplier.taminotchi_ismi}\n"
            f"Summa: {miqdor}\n"
            f"Sana: {sana}\n"
            "Ma’lumotlar to‘g‘rimi?"
        )
        confirm_buttons = [
            InlineKeyboardButton("✅ Ha", callback_data="btn_payment_confirm"),
            InlineKeyboardButton("❌ Yo‘q, bekor qilish", callback_data="btn_payment_reject")
        ]
        markup = nav_keyboard(extra=confirm_buttons)
        await update.message.reply_text(text, reply_markup=markup)
        context.user_data['current_state'] = PAYMENT_CONFIRM
        return PAYMENT_CONFIRM
    except ValueError:
        await update.message.reply_text("Sana formati noto‘g‘ri! To‘g‘ri format: DD-MM-YYYY.", reply_markup=nav_keyboard())
        return ASK_PAYMENT_DATE
from decimal import Decimal
from datetime import datetime

async def payment_confirm_handler(update, context):
    query = update.callback_query
    await query.answer()
    logger.info(
        f"payment_confirm_handler: data={query.data}, current_state={context.user_data.get('current_state')}"
    )

    if query.data == "btn_payment_confirm":
        supplier = context.user_data.get("selected_supplier")
        if not supplier:
            await query.message.reply_text(
                "Taminotchi tanlanmadi. Qaytadan boshlang.",
                reply_markup=nav_keyboard()
            )
            context.user_data['current_state'] = None
            await show_menu(update, context)
            return ConversationHandler.END

        amount = context.user_data.get("tolov_miqdori")
        sana = context.user_data.get("tolov_sanasi")
        currency = context.user_data.get('currency')

        try:
            # always convert amount to Decimal
            qoldiq = Decimal(str(amount))
            input_date = datetime.strptime(sana, "%d-%m-%Y").date()

            all_debts = await sync_to_async(list)(
                Pul_olish.objects.filter(taminotchi=supplier,status = "tolanmagan").order_by("sana")
            )
            if today>=input_date:
                if  all_debts:
                    for qarz in all_debts:
                        if qarz.currency != currency:
                            continue

                        qarz_qoldiq = qarz.umumiy_miqdor - qarz.tolangan

                        if qarz_qoldiq <= 0:
                            continue

                        if qoldiq <= qarz_qoldiq:
                            await sync_to_async(Pul_berish.objects.create)(
                                taminotchi=supplier,
                                pul_olingan=qarz,
                                sana=input_date,
                                summa=qoldiq,
                                berildi=True,
                                currency=currency
                            )
                            qarz.tolangan += qoldiq
                            if qarz.tolangan == qarz.umumiy_miqdor:
                                qarz.status = 'tolangan'
                            await sync_to_async(qarz.save)()
                            qoldiq = Decimal('0')
                            break  # break here, as qoldiq is finished
                        else:
                            # qoldiq katta, qarzni to‘liq yopamiz va qoldiqni kamaytiramiz
                            await sync_to_async(Pul_berish.objects.create)(
                                taminotchi=supplier,
                                pul_olingan=qarz,
                                sana=input_date,
                                summa=qarz_qoldiq,
                                berildi=True,
                                currency=currency
                            )
                            qarz.tolangan += qarz_qoldiq
                            qarz.status = 'tolangan'
                            await sync_to_async(qarz.save)()
                            qoldiq -= qarz_qoldiq

                    # loopdan chiqqandan keyin qoldiq qolgan bo‘lsa, umumiy pul berish yozamiz
                    if qoldiq > 0:
                        await sync_to_async(Pul_berish.objects.create)(
                            taminotchi=supplier,
                            sana=input_date,
                            summa=qoldiq,
                            berildi=True,
                            currency=currency
                        )

                else:
                        # Qarzi yopilgan bo‘lsa ham, yangi to‘lov yozishni istasangiz:
                        await sync_to_async(Pul_berish.objects.create)(
                            taminotchi=supplier,
                            
                            sana=input_date,
                            summa=qoldiq,
                            berildi = True,
                            currency = currency
                        )
            else:
                await sync_to_async(Pul_berish.objects.create)(
                            taminotchi=supplier,
                            
                            sana=input_date,
                            summa=qoldiq,
                            currency = currency
                        )
            # Qarzdorlikni hisoblash uchun metod/propertiyani sync-to-async orqali chaqirish
            umumiy_qarz = await sync_to_async(supplier.umumiy_qarz)()


            await query.message.reply_text(
                f"{amount} {currency} to‘landi.\n"
                f"{'Barcha qarzlar yopildi.' if umumiy_qarz <= 0 else f'{umumiy_qarz} so‘m hali yopilmagan.'}",
                reply_markup=nav_keyboard()
            )
            context.user_data['current_state'] = None
            await show_menu(update, context)
            return ConversationHandler.END

        except Exception as e:
            import traceback
            logger.error(f"Error in payment_confirm_handler: {e}\n{traceback.format_exc()}")
            # mana bu qatorda xabar chiqaring:
            await query.message.reply_text(
                f"Xatolik yuz berdi. Trace: {e}",  # vaqtinchalik foydalanuvchiga xato xabarini yuboradi
                reply_markup=nav_keyboard()
            )
            return PAYMENT_CONFIRM


    elif query.data == "btn_payment_reject":
        context.user_data['current_state'] = None
        await show_menu(update, context)
        return ConversationHandler.END

    elif query.data == "btn_menu":
        context.user_data['current_state'] = None
        await show_menu(update, context)
        return ConversationHandler.END

    elif query.data == "btn_back":
        await query.message.reply_text(
            "To‘lov sanasini kiriting (DD-MM-YYYY):",
            reply_markup=nav_keyboard()
        )
        context.user_data['current_state'] = ASK_PAYMENT_DATE
        return ASK_PAYMENT_DATE


# --- EXCEL HANDLERS ---
async def excel_callback_handler(update, context):
    query = update.callback_query
    await query.answer()
    if not await check_login(context, update):
        return ConversationHandler.END
    context.user_data.pop("excel_filter", None)
    context.user_data.pop("excel_start_date", None)
    keyboard = [
        [InlineKeyboardButton("Butun davr", callback_data="excel_full")],
        [InlineKeyboardButton("Vaqt oralig‘i bo‘yicha", callback_data="excel_interval")],
    ]
    markup = nav_keyboard(extra=keyboard)
    await query.message.reply_text("Ma’lumot qaysi davr uchun kerak?", reply_markup=markup)
    context.user_data['current_state'] = EXCEL_MENU
    return EXCEL_MENU

async def excel_option_handler(update, context):
    query = update.callback_query
    await query.answer()
    logger.info(f"excel_option_handler: data={query.data}, current_state={context.user_data.get('current_state')}")
    if query.data == "excel_full":
        context.user_data["excel_filter"] = ("full", None, None)
        await excel_export_handler(update, context)
        return ConversationHandler.END
    elif query.data == "excel_interval":
        await query.message.reply_text("Boshlanish sanasini kiriting (DD-MM-YYYY):", reply_markup=nav_keyboard())
        context.user_data['current_state'] = EXCEL_DATE_START
        return EXCEL_DATE_START
    elif query.data == "btn_menu":
        context.user_data['current_state'] = None
        await show_menu(update, context)
        return ConversationHandler.END
    elif query.data == "btn_back":
        supplier = context.user_data.get("selected_supplier")
        if supplier:
            return await show_supplier_info(update, context, supplier)
        context.user_data['current_state'] = None
        await show_menu(update, context)
        return ConversationHandler.END

async def excel_date_start_handler(update, context):
    text = update.message.text.strip()
    try:
        start_date = datetime.strptime(text, "%d-%m-%Y").date()
        context.user_data["excel_start_date"] = start_date
        await update.message.reply_text("Tugash sanasini kiriting (DD-MM-YYYY):", reply_markup=nav_keyboard())
        context.user_data['current_state'] = EXCEL_DATE_END
        return EXCEL_DATE_END
    except ValueError:
        await update.message.reply_text("Sana formati noto‘g‘ri! To‘g‘ri format: DD-MM-YYYY.", reply_markup=nav_keyboard())
        return EXCEL_DATE_START

async def excel_date_end_handler(update, context):
    text = update.message.text.strip()
    logger.info(f"excel_date_end_handler: text={text}, context.user_data={context.user_data}")
    try:
        end_date = datetime.strptime(text, "%d-%m-%Y").date()
        start_date = context.user_data.get("excel_start_date")
        if not start_date:
            await update.message.reply_text("Boshlanish sanasi topilmadi. Qaytadan urinib ko‘ring.", reply_markup=nav_keyboard())
            context.user_data['current_state'] = EXCEL_DATE_START
            return EXCEL_DATE_START
        if end_date < start_date:
            await update.message.reply_text("Tugash sanasi boshlanish sanasidan oldin bo‘lishi mumkin emas!", reply_markup=nav_keyboard())
            return EXCEL_DATE_END
        context.user_data["excel_filter"] = ("interval", start_date, end_date)
        await excel_export_handler(update, context, start_date=start_date, end_date=end_date)
        await show_menu(update, context)
        return ConversationHandler.END
    except ValueError:
        await update.message.reply_text("Sana formati noto‘g‘ri! To‘g‘ri format: DD-MM-YYYY.", reply_markup=nav_keyboard())
        return EXCEL_DATE_END
from app.models import get_usd_to_uzs_rate
from openpyxl.styles import Alignment, Font, PatternFill
from io import BytesIO

from decimal import Decimal, InvalidOperation


def safe_decimal(val):
    try:
        return Decimal(str(val))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal("0")

def _safe_date(s):
    try:
        return datetime.strptime(s, "%d-%m-%Y").date() if s else None
    except:
        return None

async def excel_export_handler(update, context, supplier_id=None, start_date=None, end_date=None):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl import Workbook
    import pandas as pd
    from io import BytesIO
    from datetime import datetime, date
    from decimal import Decimal
    from zoneinfo import ZoneInfo

    YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    query = getattr(update, "callback_query", None)
    if not await check_login(context, update):
        return ConversationHandler.END

    if not supplier_id:
        supplier_id = context.user_data.get("selected_supplier_id")
    if not supplier_id and query and query.data.startswith("excel_"):
        try:
            supplier_id = int(query.data.split("_")[1])
            context.user_data["selected_supplier_id"] = supplier_id
        except Exception:
            msg = "Xato: Taminotchi ID noto‘g‘ri. Qaytadan urinib ko‘ring."
            await (query.message.reply_text(msg) if query else update.message.reply_text(msg))
            return ConversationHandler.END
    if not supplier_id:
        msg = "Taminotchi tanlanmadi. Iltimos, taminotchini tanlang."
        await (query.message.reply_text(msg) if query else update.message.reply_text(msg))
        return ConversationHandler.END

    try:
        supplier = await sync_to_async(Taminotchi.objects.get)(id=supplier_id)
    except Taminotchi.DoesNotExist:
        msg = "Taminotchi topilmadi. Qaytadan urinib ko‘ring."
        await (query.message.reply_text(msg) if query else update.message.reply_text(msg))
        return ConversationHandler.END

    if not start_date or not end_date:
        excel_filter = context.user_data.get("excel_filter")
        if excel_filter and excel_filter[0] == "interval":
            start_date, end_date = excel_filter[1], excel_filter[2]
        else:
            try:
                earliest_debt = await sync_to_async(Pul_olish.objects.earliest)('sana')
                earliest_payment = await sync_to_async(Pul_berish.objects.earliest)('sana')
                latest_debt = await sync_to_async(Pul_olish.objects.latest)('sana')
                latest_payment = await sync_to_async(Pul_berish.objects.latest)('sana')
                start_date = min(earliest_debt.sana, earliest_payment.sana)
                end_date = max(latest_debt.sana, latest_payment.sana)
            except Exception:
                start_date = date.today()
                end_date = date.today()

    pul_olish_filter = {'taminotchi': supplier, 'sana__range': (start_date, end_date)}
    pul_berish_filter = {'taminotchi': supplier, 'sana__range': (start_date, end_date)}

    pul_olishlar = await sync_to_async(lambda: list(Pul_olish.objects.filter(**pul_olish_filter).order_by("sana", "id")))()
    pul_berishlar = await sync_to_async(lambda: list(Pul_berish.objects.filter(**pul_berish_filter).order_by("sana", "id")))()

    rows, rows_pb_map = [], []
    used_berish_ids = set()
    today = datetime.now(ZoneInfo("Asia/Tashkent")).date()

    for po in pul_olishlar:
        related_berish = [pb for pb in pul_berishlar if pb.pul_olingan_id == po.id]
        status_display = "to'lanmagan" if po.status == "tolanmagan" else "to'langan"
        for idx, pb in enumerate(related_berish):
            rows.append({
                "Sana": po.sana.strftime("%d-%m-%Y"),
                "Olingan qarz": float(po.umumiy_miqdor) if idx == 0 else "",
                "Valyuta olingan": po.currency if idx == 0 else "",
                "Sabab": po.sabab if idx == 0 else "",
                "Holat": status_display if idx == 0 else "",
                "Qarz to'landi": float(pb.summa),
                "Valyuta to'landi": pb.currency,
            })
            rows_pb_map.append(pb)
            used_berish_ids.add(pb.id)
        if not related_berish:
            rows.append({
                "Sana": po.sana.strftime("%d-%m-%Y"),
                "Olingan qarz": float(po.umumiy_miqdor),
                "Valyuta olingan": po.currency,
                "Sabab": po.sabab,
                "Holat": status_display,
                "Qarz to'landi": "",
                "Valyuta to'landi": "",
            })
            rows_pb_map.append(None)

    for pb in pul_berishlar:
        if pb.id not in used_berish_ids:
            rows.append({
                "Sana": pb.sana.strftime("%d-%m-%Y"),
                "Olingan qarz": "",
                "Valyuta olingan": "",
                "Sabab": "",
                "Holat": "",
                "Qarz to'landi": float(pb.summa),
                "Valyuta to'landi": pb.currency,
            })
            rows_pb_map.append(pb)

    def _safe_date(s):
        try:
            return datetime.strptime(s, "%d-%m-%Y").date() if s else None
        except:
            return None

    total_qarz_uzs = sum(
        float(row["Olingan qarz"]) for row in rows
        if row["Olingan qarz"] not in ("", None) and row["Valyuta olingan"] == "UZS"  and row.get("Holat") != "to'langan" and _safe_date(row["Sana"]) <= today
    )
    total_qarz_usd = sum(
        float(row["Olingan qarz"]) for row in rows
        if row["Olingan qarz"] not in ("", None) and row["Valyuta olingan"] == "USD" and row.get("Holat") != "to'langan" and _safe_date(row["Sana"]) <= today
    )
    total_berish_uzs = sum(
        float(row["Qarz to'landi"])
        for row, pb in zip(rows, rows_pb_map)
        if row["Qarz to'landi"] not in ("", None)
        and row["Valyuta to'landi"] == "UZS"
        and pb is not None
        and pb.berildi is True
        and pb.sana <= today
    )

    total_berish_usd = sum(
        float(row["Qarz to'landi"])
        for row, pb in zip(rows, rows_pb_map)
        if row["Qarz to'landi"] not in ("", None)
        and row["Valyuta to'landi"] == "USD"
        and pb is not None
        and pb.berildi is True
        and pb.sana <= today
    )


    summary_row = {
        "Sana": "Jami UZS",
        "Olingan qarz": total_qarz_uzs,
        "Valyuta olingan": "UZS",
        "Sabab": "",
        "Holat": "",
        "Qarz to'landi": total_berish_uzs,
        "Valyuta to'landi": "UZS",
    }
    summary_row_usd = {
        "Sana": "Jami USD",
        "Olingan qarz": total_qarz_usd,
        "Valyuta olingan": "USD",
        "Sabab": "",
        "Holat": "",
        "Qarz to'landi": total_berish_usd,
        "Valyuta to'landi": "USD",
    }
    jami_qarzdorlik = await sync_to_async(supplier.umumiy_qarz)()
    jami_row = {
        "Sana": "",
        "Olingan qarz": "",
        "Valyuta olingan": "",
        "Sabab": "",
        "Holat": "",
        "Qarz to'landi": "Jami qarzdorlik:",
        "Valyuta to'landi": f"{jami_qarzdorlik:,.2f} UZS",
    }

    rows.extend([summary_row, summary_row_usd, jami_row])

# Sort all rows by 'Sana' before converting to DataFrame
    def safe_date_key(row):
        try:
            return datetime.strptime(row["Sana"], "%d-%m-%Y").date()
        except:
            return date.max

    rows.sort(key=safe_date_key)

    df = pd.DataFrame(rows)

    with BytesIO() as output:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Sheet1'

        num_columns = len(df.columns)
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)
        cell = worksheet.cell(row=1, column=1)
        cell.value = f"Taminotchi: {supplier.taminotchi_ismi}"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_columns)
        cell2 = worksheet.cell(row=2, column=1)
        cell2.value = f"{start_date} dan {end_date} gacha" if start_date and end_date else "Barcha davrlar"
        cell2.font = Font(bold=True)
        cell2.alignment = Alignment(horizontal='center')

        for col_idx, col_name in enumerate(df.columns, start=1):
            worksheet.cell(row=3, column=col_idx, value=col_name).font = Font(bold=True)

        for row_idx, row in enumerate(df.values, start=4):
            pb_obj = rows_pb_map[row_idx - 4] if row_idx - 4 < len(rows_pb_map) else None
            for col_idx, value in enumerate(row, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                if row_idx in [len(df) + 1, len(df) + 2, len(df) + 3]:
                    cell.font = Font(bold=True)
                if pb_obj and hasattr(pb_obj, "berildi") and hasattr(pb_obj, "sana"):
                    try:
                        sana_dt = pb_obj.sana if isinstance(pb_obj.sana, date) else datetime.strptime(str(pb_obj.sana), "%Y-%m-%d").date()
                        if pb_obj.berildi is False:
                            cell.fill = YELLOW_FILL if sana_dt > today else RED_FILL
                    except:
                        pass

        for col_idx in range(1, num_columns + 1):
            max_length = len(str(df.columns[col_idx - 1]))
            for row_idx in range(3, worksheet.max_row + 1):
                try:
                    max_length = max(max_length, len(str(worksheet.cell(row=row_idx, column=col_idx).value or "")))
                except:
                    pass
            worksheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

        workbook.save(output)
        output.seek(0)
        filename = f"{supplier.taminotchi_ismi}_qarzlar.xlsx"
        if query:
            await query.message.reply_document(document=output, filename=filename)
        else:
            await update.message.reply_document(document=output, filename=filename)

    context.user_data['current_state'] = None
    if query:
        await show_menu(update, context)
    return ConversationHandler.END






# --- REPORT HANDLERS ---
async def report_option_handler(update, context):
    query = update.callback_query
    await query.answer()
    logger.info(f"report_option_handler: data={query.data}, current_state={context.user_data.get('current_state')}")
    if query.data == "report_full":
        suppliers = await sync_to_async(list)(Taminotchi.objects.all())
        if not suppliers:
            await query.message.reply_text("Hech qanday taminotchi topilmadi.", reply_markup=nav_keyboard())
            context.user_data['current_state'] = None
            await show_menu(update, context)
            return ConversationHandler.END
        try:
            earliest_debt = await sync_to_async(Pul_olish.objects.earliest)('sana')
            earliest_payment = await sync_to_async(Pul_berish.objects.earliest)('sana')
            latest_debt = await sync_to_async(Pul_olish.objects.latest)('sana')
            latest_payment = await sync_to_async(Pul_berish.objects.latest)('sana')
            start_date = min(earliest_debt.sana, earliest_payment.sana)
            end_date = max(latest_debt.sana, latest_payment.sana)
        except (Pul_olish.DoesNotExist, Pul_berish.DoesNotExist):
            start_date = today
            end_date = today
        context.user_data["report_filter"] = ("full", start_date, end_date)
        await report_export_handler(update, context, start_date=start_date, end_date=end_date)
        return ConversationHandler.END
    elif query.data == "report_interval":
        await query.message.reply_text("Boshlanish sanasini kiriting (DD-MM-YYYY):", reply_markup=nav_keyboard())
        context.user_data['current_state'] = REPORT_DATE_START
        return REPORT_DATE_START
    elif query.data == "btn_menu":
        context.user_data['current_state'] = None
        await show_menu(update, context)
        return ConversationHandler.END
    elif query.data == "btn_back":
        context.user_data['current_state'] = None
        await show_menu(update, context)
        return ConversationHandler.END

async def report_date_start_handler(update, context):
    text = update.message.text.strip()
    try:
        start_date = datetime.strptime(text, "%d-%m-%Y").date()
        context.user_data["report_start_date"] = start_date
        await update.message.reply_text("Tugash sanasini kiriting (DD-MM-YYYY):", reply_markup=nav_keyboard())
        context.user_data['current_state'] = REPORT_DATE_END
        return REPORT_DATE_END
    except ValueError:
        await update.message.reply_text("Sana formati noto‘g‘ri! To‘g‘ri format: DD-MM-YYYY.", reply_markup=nav_keyboard())
        return REPORT_DATE_START

async def report_date_end_handler(update, context):
    text = update.message.text.strip()
    logger.info(f"report_date_end_handler: text={text}, context.user_data={context.user_data}")
    try:
        end_date = datetime.strptime(text, "%d-%m-%Y").date()
        start_date = context.user_data.get("report_start_date")
        if not start_date:
            await update.message.reply_text("Boshlanish sanasi topilmadi. Qaytadan urinib ko‘ring.", reply_markup=nav_keyboard())
            context.user_data['current_state'] = REPORT_DATE_START
            return REPORT_DATE_START
        if end_date < start_date:
            await update.message.reply_text("Tugash sanasi boshlanish sanasidan oldin bo‘lishi mumkin emas!", reply_markup=nav_keyboard())
            return REPORT_DATE_END
        context.user_data["report_filter"] = ("interval", start_date, end_date)
        await report_export_handler(update, context, start_date=start_date, end_date=end_date)
        return ConversationHandler.END
    except ValueError:
        await update.message.reply_text("Sana formati noto‘g‘ri! To‘g‘ri format: DD-MM-YYYY.", reply_markup=nav_keyboard())
        return REPORT_DATE_END



# from io import BytesIO
# from datetime import date, datetime
# from zoneinfo import ZoneInfo
# import pandas as pd
# from openpyxl import Workbook
# from openpyxl.utils import get_column_letter
# from openpyxl.styles import Alignment, Font
# from openpyxl.cell.cell import MergedCell
async def report_export_handler(update, context, start_date=None, end_date=None):
    from io import BytesIO
    from datetime import date, datetime
    from zoneinfo import ZoneInfo
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.cell.cell import MergedCell

    YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    query = getattr(update, "callback_query", None)
    if not await check_login(context, update):
        return ConversationHandler.END

    suppliers = await sync_to_async(list)(Taminotchi.objects.all())
    if not suppliers:
        msg = "Hech qanday taminotchi topilmadi."
        if query:
            await query.message.reply_text(msg, reply_markup=nav_keyboard())
        else:
            await update.message.reply_text(msg, reply_markup=nav_keyboard())
        context.user_data['current_state'] = None
        await show_menu(update, context)
        return ConversationHandler.END

    if not start_date or not end_date:
        report_filter = context.user_data.get("report_filter")
        if report_filter and report_filter[0] == "interval":
            start_date = report_filter[1]
            end_date = report_filter[2]
        else:
            try:
                earliest_debt = await sync_to_async(Pul_olish.objects.earliest)('sana')
                earliest_payment = await sync_to_async(Pul_berish.objects.earliest)('sana')
                latest_debt = await sync_to_async(Pul_olish.objects.latest)('sana')
                latest_payment = await sync_to_async(Pul_berish.objects.latest)('sana')
                start_date = min(earliest_debt.sana, earliest_payment.sana)
                end_date = max(latest_debt.sana, latest_payment.sana)
            except (Pul_olish.DoesNotExist, Pul_berish.DoesNotExist):
                start_date = date.today()
                end_date = date.today()

    today = datetime.now(ZoneInfo("Asia/Tashkent")).date()
    qarz_date_to = min(today, end_date)

    # === MAIN SUMMARY SHEET (Umumiy Hisobot) ===
    # === MAIN SUMMARY SHEET (Umumiy Hisobot) ===
    summary_rows = []
    for idx, supplier in enumerate(suppliers, start=1):
        pul_olish_filter = {'taminotchi': supplier, 'sana__gte': start_date, 'sana__lte': qarz_date_to}
        pul_berish_filter = {'taminotchi': supplier, 'sana__gte': start_date, 'sana__lte': qarz_date_to}

        po_queryset = await sync_to_async(lambda: list(Pul_olish.objects.filter(**pul_olish_filter)))()
        pb_queryset = await sync_to_async(lambda: list(Pul_berish.objects.filter(**pul_berish_filter)))()

        qarz_uzs = sum(po.umumiy_miqdor for po in po_queryset if po.currency == "UZS")
        qarz_usd = sum(po.umumiy_miqdor for po in po_queryset if po.currency == "USD")
        tolangan_uzs = sum(pb.summa for pb in pb_queryset if pb.currency == "UZS")
        tolangan_usd = sum(pb.summa for pb in pb_queryset if pb.currency == "USD")

        # Convert USD to UZS (you may want to use real-time or static conversion rate)
        USD_TO_UZS = 12500
        qarz_qoldi_uzs = (qarz_uzs + qarz_usd * USD_TO_UZS) - (tolangan_uzs + tolangan_usd * USD_TO_UZS)

        summary_rows.append({
            'No.': idx,
            'Taminotchi Ismi': supplier.taminotchi_ismi,
            "Qarz olingan(UZS)": float(qarz_uzs),
            "Qarz olingan(USD)": float(qarz_usd),
            "To‘langan(UZS)": float(tolangan_uzs),
            "To‘langan(USD)": float(tolangan_usd),
            "Qarz Qoldi(UZS)": float(qarz_qoldi_uzs)
        })

    # Create DataFrame
    columns = ['No.', 'Taminotchi Ismi', 'Qarz olingan(UZS)', 'Qarz olingan(USD)', 'To‘langan(UZS)', 'To‘langan(USD)', 'Qarz Qoldi(UZS)']
    summary_df = pd.DataFrame(summary_rows, columns=columns)

    # Add total row
    if not summary_df.empty:
        total_row = {
            'No.': '',
            'Taminotchi Ismi': 'Jami:',
            'Qarz olingan(UZS)': summary_df['Qarz olingan(UZS)'].sum(),
            'Qarz olingan(USD)': summary_df['Qarz olingan(USD)'].sum(),
            'To‘langan(UZS)': summary_df['To‘langan(UZS)'].sum(),
            'To‘langan(USD)': summary_df['To‘langan(USD)'].sum(),
            'Qarz Qoldi(UZS)': summary_df['Qarz Qoldi(UZS)'].sum()
        }
        summary_df = pd.concat([summary_df, pd.DataFrame([total_row])], ignore_index=True)


    with BytesIO() as output:
        workbook = Workbook()
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])

        # === UMUMIY HISOBOT SHEET ===
        summary_sheet = workbook.create_sheet('Umumiy Hisobot')
        summary_sheet.merge_cells('A1:E1')
        summary_sheet['A1'].value = f"{start_date} dan {end_date} gacha"
        summary_sheet['A1'].alignment = Alignment(horizontal='center')
        summary_sheet['A1'].font = Font(bold=True)

        if summary_df.empty:
            summary_sheet['A2'].value = "Ma'lumotlar mavjud emas"
            summary_sheet['A2'].alignment = Alignment(horizontal='center')
            summary_sheet.merge_cells('A2:E2')
        else:
            for col_idx, col_name in enumerate(summary_df.columns, start=1):
                summary_sheet[f"{get_column_letter(col_idx)}2"] = col_name
                summary_sheet[f"{get_column_letter(col_idx)}2"].font = Font(bold=True)
            for row_idx, row in enumerate(summary_df.values, start=3):
                for col_idx, value in enumerate(row, start=1):
                    cell = summary_sheet[f"{get_column_letter(col_idx)}{row_idx}"]
                    cell.value = value
                    if row_idx == len(summary_df) + 2:
                        cell.font = Font(bold=True)

            for col_idx in range(1, summary_df.shape[1] + 1):
                col_letter = get_column_letter(col_idx)
                max_len = len(str(summary_df.columns[col_idx - 1]))

                for row_idx in range(2, summary_sheet.max_row + 1):
                    cell_value = summary_sheet[f"{col_letter}{row_idx}"].value
                    if cell_value is not None:
                        max_len = max(max_len, len(str(cell_value)))

                # Add padding to avoid cutting off longer values
                summary_sheet.column_dimensions[col_letter].width = max_len + 4

        # === HAR BIR SUPPLIER SHEET YANGILANGAN FORMATDA ===
        for supplier in suppliers:
            pul_olishlar = await sync_to_async(lambda: list(Pul_olish.objects.filter(
                taminotchi=supplier, sana__range=(start_date, end_date)).order_by("sana", "id")))()
            pul_berishlar = await sync_to_async(lambda: list(Pul_berish.objects.filter(
                taminotchi=supplier, sana__range=(start_date, end_date)).order_by("sana", "id")))()

            rows, rows_pb_map = [], []
            used_berish_ids = set()

            for po in pul_olishlar:
                related = [pb for pb in pul_berishlar if pb.pul_olingan_id == po.id]
                status = "to'lanmagan" if po.status == "tolanmagan" else "to'langan"
                for idx, pb in enumerate(related):
                    rows.append({
                        "Sana": po.sana.strftime("%d-%m-%Y"),
                        "Olingan qarz": float(po.umumiy_miqdor) if idx == 0 else "",
                        "Valyuta olingan": po.currency if idx == 0 else "",
                        "Sabab": po.sabab if idx == 0 else "",
                        "Holat": status if idx == 0 else "",
                        "Qarz to'landi": float(pb.summa),
                        "Valyuta to'landi": pb.currency
                    })
                    rows_pb_map.append(pb)
                    used_berish_ids.add(pb.id)
                if not related:
                    rows.append({
                        "Sana": po.sana.strftime("%d-%m-%Y"),
                        "Olingan qarz": float(po.umumiy_miqdor),
                        "Valyuta olingan": po.currency,
                        "Sabab": po.sabab,
                        "Holat": status,
                        "Qarz to'landi": "",
                        "Valyuta to'landi": ""
                    })
                    rows_pb_map.append(None)

            for pb in pul_berishlar:
                if pb.id not in used_berish_ids:
                    rows.append({
                        "Sana": pb.sana.strftime("%d-%m-%Y"),
                        "Olingan qarz": "",
                        "Valyuta olingan": "",
                        "Sabab": "",
                        "Holat": "",
                        "Qarz to'landi": float(pb.summa),
                        "Valyuta to'landi": pb.currency
                    })
                    rows_pb_map.append(pb)

            def _safe_date(s):
                try: return datetime.strptime(s, "%d-%m-%Y").date()
                except: return None

            qarz_uzs = sum(float(r["Olingan qarz"]) for r in rows if r["Valyuta olingan"] == "UZS" and r["Olingan qarz"] and r.get("Holat") != "to'langan" and _safe_date(r["Sana"]) <= today)
            qarz_usd = sum(float(r["Olingan qarz"]) for r in rows if r["Valyuta olingan"] == "USD" and r["Olingan qarz"] and r.get("Holat") != "to'langan" and _safe_date(r["Sana"]) <= today)
            tolandi_uzs = sum(
                float(r["Qarz to'landi"])
                for r, pb in zip(rows, rows_pb_map)
                if r["Qarz to'landi"]
                and r["Valyuta to'landi"] == "UZS"
                and pb is not None
                and pb.berildi is True
                and pb.sana <= today
            )

            tolandi_usd = sum(
                float(r["Qarz to'landi"])
                for r, pb in zip(rows, rows_pb_map)
                if r["Qarz to'landi"]
                and r["Valyuta to'landi"] == "USD"
                and pb is not None
                and pb.berildi is True
                and pb.sana <= today
            )

            jami_qarz = await sync_to_async(supplier.umumiy_qarz)()

            rows.extend([
                {"Sana": "Jami UZS", "Olingan qarz": qarz_uzs, "Valyuta olingan": "UZS", "Sabab": "", "Holat": "", "Qarz to'landi": tolandi_uzs, "Valyuta to'landi": "UZS"},
                {"Sana": "Jami USD", "Olingan qarz": qarz_usd, "Valyuta olingan": "USD", "Sabab": "", "Holat": "", "Qarz to'landi": tolandi_usd, "Valyuta to'landi": "USD"},
                {"Sana": "", "Olingan qarz": "", "Valyuta olingan": "", "Sabab": "", "Holat": "", "Qarz to'landi": "Jami qarzdorlik:", "Valyuta to'landi": f"{jami_qarz:,.2f} UZS"}
            ])

            # Sort rows by parsed 'Sana' (date string in format "%d-%m-%Y")
            def safe_date_key(row):
                try:
                    return datetime.strptime(row["Sana"], "%d-%m-%Y").date()
                except:
                    return date.max  # keep summary rows like 'Jami UZS' at the bottom

            rows.sort(key=safe_date_key)

            df = pd.DataFrame(rows)

            sheet_name = supplier.taminotchi_ismi[:31].replace('/', '_').replace('\\', '_')
            sheet = workbook.create_sheet(sheet_name)

            sheet.merge_cells('A1:G1')
            sheet['A1'].value = f"Taminotchi: {supplier.taminotchi_ismi}"
            sheet['A1'].font = Font(bold=True)
            sheet['A1'].alignment = Alignment(horizontal='center')

            sheet.merge_cells('A2:G2')
            sheet['A2'].value = f"{start_date} dan {end_date} gacha"
            sheet['A2'].font = Font(bold=True)
            sheet['A2'].alignment = Alignment(horizontal='center')

            for col_idx, col_name in enumerate(df.columns, start=1):
                sheet[f"{get_column_letter(col_idx)}3"] = col_name
                sheet[f"{get_column_letter(col_idx)}3"].font = Font(bold=True)

            for row_idx, row in enumerate(df.values, start=4):
                pb_obj = rows_pb_map[row_idx - 4] if row_idx - 4 < len(rows_pb_map) else None
                for col_idx, value in enumerate(row, start=1):
                    cell = sheet[f"{get_column_letter(col_idx)}{row_idx}"]
                    cell.value = value
                    if row_idx >= len(df) + 1:
                        cell.font = Font(bold=True)
                    if pb_obj and hasattr(pb_obj, "berildi") and hasattr(pb_obj, "sana"):
                        try:
                            sana_dt = pb_obj.sana if isinstance(pb_obj.sana, date) else datetime.strptime(str(pb_obj.sana), "%Y-%m-%d").date()
                            if pb_obj.berildi is False:
                                cell.fill = YELLOW_FILL if sana_dt > today else RED_FILL
                        except:
                            pass

            for col_idx in range(1, len(df.columns) + 1):
                col_letter = get_column_letter(col_idx)
                max_length = len(str(df.columns[col_idx - 1]))
                for row_idx in range(3, sheet.max_row + 1):
                    val = str(sheet[f"{col_letter}{row_idx}"].value or "")
                    max_length = max(max_length, len(val))
                sheet.column_dimensions[col_letter].width = max_length + 2

        workbook.save(output)
        output.seek(0)
        filename = f"Umumiy_Hisobot_{start_date}_to_{end_date}.xlsx"
        if query:
            await query.message.reply_document(document=output, filename=filename)
        else:
            await update.message.reply_document(document=output, filename=filename)

    context.user_data['current_state'] = None
    await show_menu(update, context)
    return ConversationHandler.END

#NOTIFICATION
# -------------------- SUPPLIER LIST --------------------
async def rejala_supplier_list(update, context):
    from telegram import InlineKeyboardButton, InlineKeyboardMarkup

    page = int(context.user_data.get('rejala_supplier_page', 1))
    all_suppliers = await sync_to_async(list)(
        Pul_berish.objects.filter(berildi=False)
        .values('taminotchi__id', 'taminotchi__taminotchi_ismi')
        .distinct().order_by('taminotchi__taminotchi_ismi')
    )

    if not all_suppliers:
        await update.callback_query.message.reply_text("Hech qanday rejalashtirilgan to‘lov yo‘q.")
        return ConversationHandler.END

    total = len(all_suppliers)
    max_page = (total - 1) // PAGINATION_SIZE + 1
    page = max(1, min(page, max_page))
    start, end = (page - 1) * PAGINATION_SIZE, page * PAGINATION_SIZE
    page_suppliers = all_suppliers[start:end]

    context.user_data['rejala_supplier_list'] = all_suppliers
    context.user_data['rejala_supplier_page'] = page

    text = "Tanlang (raqam yoki ID):\n"
    for idx, sup in enumerate(page_suppliers, start=1):
        text += f"{idx}. {sup['taminotchi__taminotchi_ismi']} (ID: {sup['taminotchi__id']})\n"
    text += f"\nSahifa: {page}/{max_page}\n"
    text += "Raqam yoki ID ni yuboring:"

    nav_buttons = []
    if page > 1:
        nav_buttons.append(InlineKeyboardButton("⬅️ Oldingi", callback_data="prev_suppliers"))
    if page < max_page:
        nav_buttons.append(InlineKeyboardButton("Keyingi ➡️", callback_data="next_suppliers"))

    # Add Ortga button
    markup = InlineKeyboardMarkup([
        nav_buttons,
        [InlineKeyboardButton("🔙 Ortga", callback_data="btn_menu")]
    ] if nav_buttons else [
        [InlineKeyboardButton("🔙 Ortga", callback_data="btn_menu")]
    ])

    await update.callback_query.message.reply_text(text, reply_markup=markup)
    return REJALA_SUPPLIER_LIST



async def rejala_supplier_choose(update, context):
    user_input = update.message.text.strip()
    page = context.user_data.get('rejala_supplier_page', 1)
    all_suppliers = context.user_data.get('rejala_supplier_list', [])
    start, end = (page-1)*PAGINATION_SIZE, page*PAGINATION_SIZE
    page_suppliers = all_suppliers[start:end]

    chosen = None
    if user_input.isdigit():
        idx = int(user_input)
        if 1 <= idx <= len(page_suppliers):
            chosen = page_suppliers[idx-1]
        else:
            for sup in all_suppliers:
                if str(sup['taminotchi__id']) == user_input:
                    chosen = sup
                    break
    if not chosen:
        await update.message.reply_text("To‘g‘ri raqam yoki ID yuboring.")
        return REJALA_SUPPLIER_LIST

    context.user_data['chosen_rejala_supplier'] = chosen
    return await rejala_payment_list(update, context)


async def rejala_payment_list(update, context):
    from telegram import InlineKeyboardButton, InlineKeyboardMarkup

    supplier = context.user_data.get('chosen_rejala_supplier')
    payments = await sync_to_async(list)(
        Pul_berish.objects.filter(taminotchi_id=supplier['taminotchi__id'], berildi=False).order_by('sana')
    )
    if not payments:
        await update.message.reply_text("Ushbu taminotchida rejalashtirilgan to‘lov yo‘q.")
        return ConversationHandler.END

    context.user_data['rejala_payment_list'] = [p.id for p in payments]

    text = f"{supplier['taminotchi__taminotchi_ismi']} uchun rejalashtirilgan to‘lovlar:\n"
    for idx, p in enumerate(payments, start=1):
        text += f"{idx}. {p.sana.strftime('%d-%m-%Y')} | {p.summa} so‘m (ID: {p.id})\n"
    text += "\nRaqam yoki ID yuboring:"

    # Ortga button
    markup = InlineKeyboardMarkup([
        [InlineKeyboardButton("🔙 Ortga", callback_data="btn_menu")]
    ])

    await update.message.reply_text(text, reply_markup=markup)
    return REJALA_PAYMENT_LIST


async def rejala_payment_choose(update, context):
    user_input = update.message.text.strip()
    payment_ids = context.user_data.get('rejala_payment_list', [])
    payments = await sync_to_async(list)(Pul_berish.objects.filter(id__in=payment_ids).order_by('sana'))
    payment = None
    if user_input.isdigit():
        idx = int(user_input)
        if 1 <= idx <= len(payments):
            payment = payments[idx-1]
        else:
            for p in payments:
                if str(p.id) == user_input:
                    payment = p
                    break
    if not payment:
        await update.message.reply_text("To‘g‘ri raqam yoki ID yuboring.")
        return REJALA_PAYMENT_LIST
    context.user_data['chosen_rejala_payment'] = payment.id
    return await show_payment_actions(update, context, payment)

async def show_payment_actions(update, context, payment):
    taminotchi = await sync_to_async(lambda: payment.taminotchi)()
    text = f"Taminotchi: {taminotchi}\nSana: {payment.sana.strftime('%d-%m-%Y')}\nSumma: {payment.summa} so‘m"

    buttons = [
        [InlineKeyboardButton("✅ To‘landi", callback_data=f"payment_done_{payment.id}")],
        [InlineKeyboardButton("🗑 Bekor qilish", callback_data=f"payment_delete_{payment.id}")],
        [InlineKeyboardButton("✏️ Keyinroq to'lash", callback_data=f"payment_edit_{payment.id}")],
        [InlineKeyboardButton("⬅️ Orqaga", callback_data="btn_rejala_back")]
    ]
    await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(buttons))
    return REJALA_PAYMENT_ACTION

from django.db.models import F
from decimal import Decimal
from datetime import datetime
from zoneinfo import ZoneInfo

# --- Payment Actions ---
async def payment_done_handler(update, context):
    payment_id = int(update.callback_query.data.split("_")[-1])
    payment = await sync_to_async(Pul_berish.objects.get)(id=payment_id)
    payment.berildi = True
    pul_olingan = await sync_to_async(
        lambda: list(
            Pul_olish.objects.filter(
                tolangan__lt=F('umumiy_miqdor'),
                status='tolanmagan'
            ).order_by('sana')
        )
    )()

    qoldiq = payment.summa
    for i in pul_olingan:
        if qoldiq <= 0:
            break
        kerak_sum = i.umumiy_miqdor - i.tolangan
        if qoldiq >= kerak_sum:
            i.tolangan += kerak_sum
            qoldiq -= kerak_sum
            if i.tolangan >= i.umumiy_miqdor:
                i.status = "tolangan"
        elif qoldiq < kerak_sum:
            i.tolangan += qoldiq
            qoldiq = Decimal('0')
        await sync_to_async(i.save)()

    payment.sana = datetime.now(ZoneInfo("Asia/Tashkent")).date()
    await sync_to_async(payment.save)()
    await update.callback_query.message.reply_text("To‘lov to‘landi sifatida belgilandi.")
    await show_menu(update, context)
    return ConversationHandler.END

async def payment_delete_handler(update, context):
    payment_id = int(update.callback_query.data.split("_")[-1])
    await sync_to_async(Pul_berish.objects.filter(id=payment_id).delete)()
    await update.callback_query.message.reply_text("To‘lov o‘chirildi.")
    await show_menu(update, context)
    return ConversationHandler.END

async def payment_edit_amount(update, context):
    payment_id = int(update.callback_query.data.split("_")[-1])
    context.user_data['edit_payment_id'] = payment_id
    await update.callback_query.message.reply_text("Yangi miqdorni yuboring:")
    return PAYMENT_EDIT_AMOUNT

async def payment_edit_amount_save(update, context):
    try:
        amount = float(update.message.text.strip())
    except:
        await update.message.reply_text("Raqam yuboring.")
        return PAYMENT_EDIT_AMOUNT
    payment_id = context.user_data['edit_payment_id']
    payment = await sync_to_async(Pul_berish.objects.get)(id=payment_id)
    payment.summa = amount
    await sync_to_async(payment.save)()
    await update.message.reply_text("Miqdor o‘zgartirildi. Sana o‘zgartirish uchun sanani yuboring (YYYY-MM-DD):")
    await show_menu(update, context)
    return ConversationHandler.END

async def payment_edit_date_save(update, context):
    try:
        sana = datetime.strptime(update.message.text.strip(), "%Y-%m-%d").date()
    except:
        await update.message.reply_text("To‘g‘ri sana yuboring (YYYY-MM-DD).")
        return PAYMENT_EDIT_DATE
    payment_id = context.user_data['edit_payment_id']
    payment = await sync_to_async(Pul_berish.objects.get)(id=payment_id)
    payment.sana = sana
    await sync_to_async(payment.save)()
    await update.message.reply_text("Sana o‘zgartirildi.")
    await show_menu(update, context)
    return ConversationHandler.END

async def rejala_payment_back(update, context):
    await show_menu(update, context)
    return ConversationHandler.END

async def rejala_payment_page(update, context):
    query = update.callback_query
    action = query.data
    page = context.user_data.get('rejala_page', 1)
    if action == "prev_rejala":
        page -= 1
    elif action == "next_rejala":
        page += 1
    context.user_data['rejala_page'] = max(page, 1)
    return await rejala_payment_list(update, context)

async def rejala_supplier_page(update, context):
    if update.callback_query.data == "prev_suppliers":
        context.user_data['rejala_supplier_page'] = context.user_data.get('rejala_supplier_page', 1) - 1
    else:
        context.user_data['rejala_supplier_page'] = context.user_data.get('rejala_supplier_page', 1) + 1
    return await rejala_supplier_list(update, context)


#XABARLAR
async def xabar_payment_list(update, context):
    from zoneinfo import ZoneInfo
    from datetime import datetime
    from telegram import InlineKeyboardMarkup, InlineKeyboardButton

    today = datetime.now(ZoneInfo("Asia/Tashkent")).date()

    context.user_data['xabar_page'] = 1
    context.user_data['xabar_payment_objs'] = []

    all_payments = await sync_to_async(list)(
        Pul_berish.objects.filter(berildi=False, sana__lte=today)
        .select_related('taminotchi')
        .order_by('sana')
    )

    if not all_payments:
        await update.callback_query.message.reply_text("Bugungi va o'tgan kunlarda to‘lanmaganlar yo‘q.")
        return ConversationHandler.END

    page = int(context.user_data.get('xabar_page', 1))
    total = len(all_payments)
    max_page = (total - 1) // PAGINATION_SIZE + 1
    page = max(1, min(page, max_page))
    start, end = (page - 1) * PAGINATION_SIZE, page * PAGINATION_SIZE
    page_payments = all_payments[start:end]
    context.user_data['xabar_payment_objs'] = all_payments
    context.user_data['xabar_page'] = page

    text = "To‘lanmaganlar ro‘yxati (raqam yoki ID):\n"
    for idx, p in enumerate(page_payments, start=1):
        text += f"{idx}. {p.taminotchi} | {p.sana.strftime('%d-%m-%Y')} | {p.summa} so‘m (ID: {p.id})\n"
    text += f"\nSahifa: {page}/{max_page}\n"
    text += "Raqam yoki ID yuboring:"

    # Pagination and back buttons
    nav_buttons = []
    if page > 1:
        nav_buttons.append(InlineKeyboardButton("⬅️ Oldingi", callback_data="prev_xabars"))
    if page < max_page:
        nav_buttons.append(InlineKeyboardButton("Keyingi ➡️", callback_data="next_xabars"))

    # "Ortga" always shown
    markup = InlineKeyboardMarkup([
        nav_buttons,
        [InlineKeyboardButton("🔙 Ortga", callback_data="btn_menu")]
    ] if nav_buttons else [
        [InlineKeyboardButton("🔙 Ortga", callback_data="btn_menu")]
    ])

    await update.callback_query.message.reply_text(text, reply_markup=markup)
    return XABAR_PAYMENT_LIST


async def xabar_payment_choose(update, context):
    user_input = update.message.text.strip()
    page = context.user_data.get('xabar_page', 1)
    all_payments = context.user_data.get('xabar_payment_objs', [])

    start, end = (page-1)*PAGINATION_SIZE, page*PAGINATION_SIZE
    page_payments = all_payments[start:end]
    chosen = None
    if user_input.isdigit():
        idx = int(user_input)
        if 1 <= idx <= len(page_payments):
            chosen = page_payments[idx-1]
        else:
            for p in all_payments:
                if str(p.id) == user_input:
                    chosen = p
                    break
    if not chosen:
        await update.message.reply_text("To‘g‘ri raqam yoki ID yuboring.")
        return XABAR_PAYMENT_LIST
    context.user_data['chosen_xabar_payment'] = chosen.id
    await show_payment_actions(update, context, chosen)  # <-- Faqat await, return emas
    return XABAR_PAYMENT_ACTION  # <-- STATE return qilinsin!


async def xabar_payment_page(update, context):
    if update.callback_query.data == "prev_xabars":
        context.user_data['xabar_page'] = context.user_data.get('xabar_page', 1) - 1
    else:
        context.user_data['xabar_page'] = context.user_data.get('xabar_page', 1) + 1
    return await xabar_payment_list(update, context)

# --- XABARLAR PAYMENT ACTIONS (just copy your existing payment_done_handler etc, but for xabarlar) ---

async def xabar_payment_done_handler(update, context):
    payment_id = int(update.callback_query.data.split("_")[-1])
    payment = await sync_to_async(Pul_berish.objects.get)(id=payment_id)
    payment.berildi = True
    # You can add your payment tolangan logic if needed
    payment.sana = datetime.now(ZoneInfo("Asia/Tashkent")).date()
    await sync_to_async(payment.save)()
    await update.callback_query.message.reply_text("To‘lov to‘landi sifatida belgilandi.")
    await show_menu(update, context)
    return ConversationHandler.END

async def xabar_payment_delete_handler(update, context):
    payment_id = int(update.callback_query.data.split("_")[-1])
    await sync_to_async(Pul_berish.objects.filter(id=payment_id).delete)()
    await update.callback_query.message.reply_text("To‘lov o‘chirildi.")
    await show_menu(update, context)
    return ConversationHandler.END

async def xabar_payment_edit_amount(update, context):
    payment_id = int(update.callback_query.data.split("_")[-1])
    context.user_data['edit_payment_id'] = payment_id
    await update.callback_query.message.reply_text("Yangi miqdorni yuboring:")
    return PAYMENT_EDIT_AMOUNT

async def xabar_payment_edit_amount_save(update, context):
    try:
        amount = float(update.message.text.strip())
    except:
        await update.message.reply_text("Raqam yuboring.")
        return PAYMENT_EDIT_AMOUNT
    payment_id = context.user_data['edit_payment_id']
    payment = await sync_to_async(Pul_berish.objects.get)(id=payment_id)
    payment.summa = amount
    await sync_to_async(payment.save)()
    await update.message.reply_text("Miqdor o‘zgartirildi. Sana o‘zgartirish uchun sanani yuboring (YYYY-MM-DD):")
    await show_menu(update, context)
    return ConversationHandler.END

async def xabar_payment_edit_date_save(update, context):
    try:
        sana = datetime.strptime(update.message.text.strip(), "%Y-%m-%d").date()
    except:
        await update.message.reply_text("To‘g‘ri sana yuboring (YYYY-MM-DD).")
        return PAYMENT_EDIT_DATE
    payment_id = context.user_data['edit_payment_id']
    payment = await sync_to_async(Pul_berish.objects.get)(id=payment_id)
    payment.sana = sana
    await sync_to_async(payment.save)()
    await update.message.reply_text("Sana o‘zgartirildi.")
    await show_menu(update, context)
    return ConversationHandler.END

async def xabar_payment_back(update, context):
    await show_menu(update, context)
    return ConversationHandler.END

# --- XABARLAR PAGINATION REMAINS THE SAME ---

async def xabar_payment_page(update, context):
    if update.callback_query.data == "prev_xabars":
        context.user_data['xabar_page'] = context.user_data.get('xabar_page', 1) - 1
    else:
        context.user_data['xabar_page'] = context.user_data.get('xabar_page', 1) + 1
    return await xabar_payment_list(update, context)


# --- HANDLER REGISTRATION ---
app = ApplicationBuilder().token("8002360442:AAHrdMm2lWTfKXRWkoeSmYR97sgNMAJtuM8").build()
rejalashtirilgan_tolovlar_conv = ConversationHandler(
    entry_points=[CallbackQueryHandler(rejala_supplier_list, pattern="^btn_rejala_tolovlar$")],
    states={
        REJALA_SUPPLIER_LIST: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, rejala_supplier_choose),
            CallbackQueryHandler(rejala_supplier_page, pattern="^(prev_suppliers|next_suppliers)$"),
            CallbackQueryHandler(show_menu, pattern="^btn_menu$")
        ],
        REJALA_PAYMENT_LIST: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, rejala_payment_choose),
            CallbackQueryHandler(rejala_payment_page, pattern="^(prev_payments|next_payments)$"),
            CallbackQueryHandler(rejala_supplier_list, pattern="^btn_rejala_tolovlar$"),
            CallbackQueryHandler(show_menu, pattern="^btn_menu$")
        ],
        REJALA_PAYMENT_ACTION: [
            CallbackQueryHandler(payment_done_handler, pattern=r"^payment_done_\d+$"),
            CallbackQueryHandler(payment_delete_handler, pattern=r"^payment_delete_\d+$"),
            CallbackQueryHandler(payment_edit_amount, pattern=r"^payment_edit_\d+$"),
            CallbackQueryHandler(rejala_payment_back, pattern="^btn_rejala_back$"),
            CallbackQueryHandler(show_menu, pattern="^btn_menu$")
        ],
        PAYMENT_EDIT_AMOUNT: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, payment_edit_amount_save),
            CallbackQueryHandler(rejala_payment_back, pattern="^btn_rejala_back$")
        ],
        PAYMENT_EDIT_DATE: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, payment_edit_date_save),
            CallbackQueryHandler(rejala_payment_back, pattern="^btn_rejala_back$")
        ]
    },
    fallbacks=[CallbackQueryHandler(show_menu, pattern="^btn_menu$")],
    per_user=True
)
xabarlar_conv = ConversationHandler(
    entry_points=[CallbackQueryHandler(xabar_payment_list, pattern="^btn_xabarlar$")],
    states={
        XABAR_PAYMENT_LIST: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, xabar_payment_choose),
            CallbackQueryHandler(xabar_payment_page, pattern="^(prev_xabars|next_xabars)$"),
            CallbackQueryHandler(show_menu, pattern="^btn_menu$")
        ],
        XABAR_PAYMENT_ACTION: [    # BU QATOR BO‘LISHI SHART!
            CallbackQueryHandler(payment_done_handler, pattern=r"^payment_done_\d+$"),
            CallbackQueryHandler(payment_delete_handler, pattern=r"^payment_delete_\d+$"),
            CallbackQueryHandler(payment_edit_amount, pattern=r"^payment_edit_\d+$"),
            CallbackQueryHandler(xabar_payment_list, pattern="^btn_xabarlar$"),
            CallbackQueryHandler(show_menu, pattern="^btn_menu$")
        ],
        PAYMENT_EDIT_AMOUNT: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, payment_edit_amount_save),
            CallbackQueryHandler(xabar_payment_list, pattern="^btn_xabarlar$")
        ],
        PAYMENT_EDIT_DATE: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, payment_edit_date_save),
            CallbackQueryHandler(xabar_payment_list, pattern="^btn_xabarlar$")
        ]
    },
    fallbacks=[CallbackQueryHandler(show_menu, pattern="^btn_menu$")],
    per_user=True
)




# Login conversation
login_conv = ConversationHandler(
    entry_points=[CommandHandler("start", start)],
    states={
        ASK_LOGIN: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, login_handler),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
    },
    fallbacks=[
        CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
    ],
    per_user=True,
)

# Supplier addition conversation
supplier_conv = ConversationHandler(
    entry_points=[CallbackQueryHandler(inline_callback, pattern="^btn_taminotchi$")],
    states={
        SUPPLIER_NAME: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, supplier_name),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        SUPPLIER_PHONE: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, supplier_phone),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
    },
    fallbacks=[
        CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
    ],
    per_user=True,
)

# Main conversation (suppliers, debts, payments, Excel)
qarz_conv = ConversationHandler(
    entry_points=[CallbackQueryHandler(taminotchilar_callback, pattern="^btn_taminotchilar$")],
    states={
        CHOOSE_SUPPLIER: [
            CallbackQueryHandler(taminotchilar_callback, pattern="^(prev_suppliers|next_suppliers|add_supplier|find_supplier)$"),
            MessageHandler(filters.TEXT & ~filters.COMMAND, choose_supplier),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        SUPPLIER_NAME: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, supplier_name),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        SUPPLIER_PHONE: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, supplier_phone),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        FIND_SUPPLIER: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, find_supplier),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        SUPPLIER_ACTION: [
            CallbackQueryHandler(excel_callback_handler, pattern=r"^excel_\d+$"),
            CallbackQueryHandler(supplier_action_handler, pattern="^(btn_qarz_bolish|btn_qarzni_tolash|btn_back|btn_menu)$"),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        PAYMENT_CURRENCY: [
            CallbackQueryHandler(ask_currency, pattern="^(currency_sum|currency_usd)$"),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$"),
        ],
        ASK_AMOUNT: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, ask_amount),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        ASK_DATE: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, ask_date),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        ASK_REASON: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, ask_reason),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        CONFIRMATION: [
            CallbackQueryHandler(confirmation_handler, pattern="^(btn_confirm|btn_reject|btn_back|btn_menu)$")
        ],
        ASK_PAYMENT_CURRENCY: [
            CallbackQueryHandler(ask_payment_currency, pattern="^(currency_sum|currency_usd)$"),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$"),
        ],

        ASK_PAYMENT_AMOUNT: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, ask_payment_amount),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        ASK_PAYMENT_DATE: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, ask_payment_date),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        PAYMENT_CONFIRM: [
            CallbackQueryHandler(payment_confirm_handler, pattern="^(btn_payment_confirm|btn_payment_reject|btn_back|btn_menu)$")
        ],
        EXCEL_MENU: [
            CallbackQueryHandler(excel_option_handler, pattern="^(excel_full|excel_interval|btn_back|btn_menu)$")
        ],
        EXCEL_DATE_START: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, excel_date_start_handler),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        EXCEL_DATE_END: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, excel_date_end_handler),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
    },
    fallbacks=[
        CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
    ],
    per_user=True
)

# Report conversation
report_conv = ConversationHandler(
    entry_points=[CallbackQueryHandler(inline_callback, pattern="^btn_hisobot$")],
    states={
        REPORT_MENU: [
            CallbackQueryHandler(report_option_handler, pattern="^(report_full|report_interval|btn_back|btn_menu)$")
        ],
        REPORT_DATE_START: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, report_date_start_handler),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        REPORT_DATE_END: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, report_date_end_handler),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
    },
    fallbacks=[
        CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
    ],
    per_user=True
)

# NEW: Ishchilar conversation handler

ishchilar_conv = ConversationHandler(
    entry_points=[
        CallbackQueryHandler(show_ishchilar, pattern="^btn_ishchilar$"),
        CallbackQueryHandler(ishchilar_callback, pattern="^(prev_ishchilar|next_ishchilar|find_ishchi|add_ishchi)$")
    ],
    states={
        CHOOSE_ISHCHI: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, choose_ishchi),
            CallbackQueryHandler(ishchilar_callback, pattern="^(prev_ishchilar|next_ishchilar|find_ishchi|add_ishchi)$"),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        ISHCHI_NAME: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, ishchi_name),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        ISHCHI_SURNAME: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, ishchi_surname),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        ISHCHI_PHONE: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, ishchi_phone),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        ISHCHI_PERMISSIONS: [
            CallbackQueryHandler(ishchi_permissions, pattern="^(perm_yes|perm_no)$"),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        ISHCHI_CONFIRM: [
            CallbackQueryHandler(ishchi_confirm, pattern="^(confirm_yes|confirm_no)$"),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        FIND_ISHCHI: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, find_ishchi),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        ISHCHI_ACTION: [
            CallbackQueryHandler(ishchi_action_handler, pattern="^(delete_ishchi|edit_ishchi)$"),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        ISHCHI_DELETE_CONFIRM: [
            CallbackQueryHandler(ishchi_delete_confirm_handler, pattern="^(confirm_delete|cancel_delete)$"),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        ISHCHI_EDIT_NAME: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, ishchi_edit_name),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        ISHCHI_EDIT_SURNAME: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, ishchi_edit_surname),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        ISHCHI_EDIT_PHONE: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, ishchi_edit_phone),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        ISHCHI_EDIT_PERMISSIONS: [
            CallbackQueryHandler(ishchi_edit_permissions, pattern="^(perm_yes|perm_no)$"),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        ISHCHI_EDIT_CONFIRM: [
            CallbackQueryHandler(ishchi_edit_confirm, pattern="^(confirm_yes|confirm_no)$"),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
    },
    fallbacks=[
        CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
    ],
    per_user=True
)

# NEW: Harajatlar conversation handler
harajat_conv = ConversationHandler(
    entry_points=[CallbackQueryHandler(inline_callback, pattern="^btn_harajatlar$")],
    states={
        HARAJAT_MENU: [
            CallbackQueryHandler(harajat_menu_handler, pattern="^(harajat_ozimga|harajat_ishxonaga|harajat_ishchilarga|harajat_list|btn_back)$"),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_menu)$")
        ],
        HARAJAT_TYPE: [
            CallbackQueryHandler(harajat_type_handler, pattern="^(harajat_ozimga|harajat_ishxonaga|harajat_ishchilarga)$"),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        HARAJAT_AMOUNT: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, harajat_amount),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        HARAJAT_DATE: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, harajat_date),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        HARAJAT_REASON: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, harajat_reason),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        HARAJAT_CONFIRMATION: [
            CallbackQueryHandler(harajat_confirmation, pattern="^(confirm_yes|confirm_no)$"),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        CHOOSE_ISHCHI_EXPENSE: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, choose_ishchi_expense),
            CallbackQueryHandler(harajat_callback, pattern="^(prev_ishchilar_harajat|next_ishchilar_harajat)$"),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        HARAJAT_LIST: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, choose_harajat_by_number),
            CallbackQueryHandler(harajat_list_callback, pattern="^(prev_harajat|next_harajat|harajat_action_\d+|harajat_excel)$"),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        HARAJAT_ACTION: [
            CallbackQueryHandler(harajat_action_handler, pattern="^(delete_harajat|edit_harajat)$"),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        HARAJAT_DELETE_CONFIRM: [
            CallbackQueryHandler(harajat_delete_confirm, pattern="^(confirm_delete|cancel_delete)$"),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        HARAJAT_EDIT_AMOUNT: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, harajat_edit_amount),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        HARAJAT_EDIT_DATE: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, harajat_edit_date),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        HARAJAT_EDIT_REASON: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, harajat_edit_reason),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        HARAJAT_EDIT_CONFIRM: [
            CallbackQueryHandler(harajat_edit_confirm, pattern="^(confirm_yes|confirm_no)$"),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        HARAJAT_EXCEL_MENU: [
            CallbackQueryHandler(harajat_excel_menu, pattern="^(excel_full|excel_interval)$"),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        HARAJAT_EXCEL_DATE_START: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, harajat_excel_date_start),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
        HARAJAT_EXCEL_DATE_END: [
            MessageHandler(filters.TEXT & ~filters.COMMAND, harajat_excel_date_end),
            CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
        ],
    },
    fallbacks=[
        CallbackQueryHandler(navigation_callback, pattern="^(btn_back|btn_menu)$")
    ],
    per_user=True
)




from telegram import BotCommand
async def set_bot_commands(application):
    commands = [
        BotCommand("menu", "Menu"),
        BotCommand("start", "Shaxsiy hisobga kirish"),
        BotCommand("logout", "Hisobdan chiqish"),
      
        # Add more commands here
    ]
    await application.bot.set_my_commands(commands)


# Entry pointlar


# Register handlers
if __name__ == '__main__':
    import django, os
    os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'your_project.settings')
    django.setup()
    set_bot_commands(app)
    app.job_queue.run_repeating(check_due_payments, interval=86400, first=0)  # Run every day
    app.add_handler(login_conv)
    app.add_handler(supplier_conv)
    app.add_handler(qarz_conv)
    app.add_handler(report_conv)
    app.add_handler(ishchilar_conv)
    app.add_handler(harajat_conv)
    app.add_handler(rejalashtirilgan_tolovlar_conv)
    app.add_handler(xabarlar_conv)
    app.add_handler(CommandHandler("menu", show_menu))
    app.add_handler(CommandHandler("logout", logout))
    app.run_polling()